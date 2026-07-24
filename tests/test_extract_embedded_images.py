"""extract_embedded_images 單元契約測試（純檔案解析，不碰 DB）。

對應維運修復 B-1：大 xlsx 匯入卡死。根因為 extract_embedded_images 為抓內嵌圖
以 read_only=False 全載整個 workbook（解析全部 drawing/樣式/公式），53.8MB 檔會 hang。
本測試鎖住兩件事，作為改寫成 zipfile 直讀後不得回歸的基準：

1. 對映正確性（最高風險）：回傳 {1-based 列號: (第一張圖 bytes, 該列圖片張數)}
   必須與錨點列一致；同列多張回第一張 + 張數；多列各自對應不錯位。
2. 效能修復本身：不得再以 read_only=False 全載 workbook（改用 zipfile 直讀），
   直接斷言新版完全不呼叫 openpyxl.load_workbook。
3. 容錯：非 zip／損毀／無圖／結構不支援 → 回空 dict，匯入照常。
"""
from __future__ import annotations

import io
import struct
import zlib
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

from backend.app.importers import wips_importer
from backend.app.importers.wips_importer import extract_embedded_images


# 與既有兩份代表圖回歸測試相同的最小合法 PNG 產生器：seed 讓每張圖 bytes 互異，
# 才能精確驗證「哪張圖落在哪一列」不錯位。
_PNG_HEAD = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108000000003a7e9b55"
)


def _png_bytes(seed: int) -> bytes:
    """產生內容互異的小張測試 PNG（seed 只影響 IDAT，仍是合法 PNG）。"""
    raw = zlib.compress(bytes([0, seed & 0xFF]))
    idat = struct.pack(">I", len(raw)) + b"IDAT" + raw
    idat += struct.pack(">I", zlib.crc32(b"IDAT" + raw) & 0xFFFFFFFF)
    iend = struct.pack(">I", 0) + b"IEND" + struct.pack(">I", zlib.crc32(b"IEND") & 0xFFFFFFFF)
    return _PNG_HEAD + idat + iend


def _build_xlsx(path: Path, sheet_title: str, n_rows: int, images: list[tuple[int, bytes]]) -> None:
    """造測試 xlsx：n_rows 筆資料列，images 為 (0-based anchor row, 圖 bytes)。

    與真檔一致：主附图 儲存格值為單一空白，圖片為錨在資料列的浮動物件。
    anchor "A{1-based}" 由 openpyxl 序列化為 drawing XML 的 <from><row>{0-based}</row>。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(["主附图", "申请号", "标题"])
    for i in range(1, n_rows + 1):
        ws.append([" ", f"TW-{i:03d}", f"標題{i}"])
    for anchor_row, blob in images:
        img = XLImage(io.BytesIO(blob))
        img.anchor = f"A{anchor_row + 1}"
        ws.add_image(img)
    wb.save(path)


class ExtractEmbeddedImagesTests(unittest.TestCase):
    def setUp(self):
        import os

        self._tmp = Path(os.environ.get("TEMP", "/tmp")) / "extract_embed_tests"
        self._tmp.mkdir(parents=True, exist_ok=True)

    def test_mapping_multi_row_and_multi_image_on_one_row(self):
        """多列、某列兩張圖：回 {列號:(第一張,張數)}，對映與錨點一致、不錯位。

        列 2（0-based row 1）一張、列 3（row 2）兩張、列 4（row 3）一張。
        期望：{2:(a,1), 3:(b,2), 4:(d,1)}，第 3 列取第一張 b、記 2 張。
        """
        a, b, c, d = _png_bytes(11), _png_bytes(22), _png_bytes(33), _png_bytes(44)
        path = self._tmp / "multi.xlsx"
        _build_xlsx(
            path,
            "download",
            n_rows=3,
            images=[(1, a), (2, b), (2, c), (3, d)],
        )
        result = extract_embedded_images(path, "download")
        self.assertEqual(result[2], (a, 1))
        self.assertEqual(result[3], (b, 2), "同列多圖取第一張並記張數")
        self.assertEqual(result[4], (d, 1))
        self.assertEqual(set(result.keys()), {2, 3, 4})

    def test_does_not_fully_load_workbook(self):
        """效能修復核心：不得再以 read_only=False 全載 workbook（B-1 根因）。

        改寫成 zipfile 直讀後，extract_embedded_images 完全不需 openpyxl.load_workbook。
        直接斷言未呼叫 load_workbook（若沿用舊寫法此測試必失敗）。
        """
        path = self._tmp / "perf.xlsx"
        _build_xlsx(path, "download", n_rows=2, images=[(1, _png_bytes(1)), (2, _png_bytes(2))])

        calls: list[dict] = []
        original = wips_importer.load_workbook

        def spy(*args, **kwargs):
            calls.append(kwargs)
            return original(*args, **kwargs)

        wips_importer.load_workbook = spy
        try:
            result = extract_embedded_images(path, "download")
        finally:
            wips_importer.load_workbook = original

        self.assertEqual(len(result), 2, "改寫後仍須正確取得兩列圖")
        self.assertEqual(
            calls,
            [],
            f"extract_embedded_images 不得再呼叫 load_workbook（B-1 全載根因），實際呼叫 {calls}",
        )

    def test_no_images_returns_empty(self):
        """無內嵌圖：回空 dict，匯入照常。"""
        path = self._tmp / "no_images.xlsx"
        _build_xlsx(path, "download", n_rows=2, images=[])
        self.assertEqual(extract_embedded_images(path, "download"), {})

    def test_corrupt_or_non_zip_returns_empty(self):
        """非 zip／損毀檔：容錯回空 dict，不得讓整批匯入失敗。"""
        path = self._tmp / "broken.xlsx"
        path.write_bytes(b"this is not a zip file")
        self.assertEqual(extract_embedded_images(path, "download"), {})

    def test_unknown_sheet_returns_empty(self):
        """指定工作表不存在：回空 dict。"""
        path = self._tmp / "sheet_missing.xlsx"
        _build_xlsx(path, "download", n_rows=1, images=[(1, _png_bytes(9))])
        self.assertEqual(extract_embedded_images(path, "not_there"), {})


if __name__ == "__main__":
    unittest.main()
