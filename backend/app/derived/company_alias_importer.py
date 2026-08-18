from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from backend.app.mappings.wips import PEOPLE_FIELD_COLUMNS
from backend.app.transforms.text import clean_text


#: 母體範圍豁免（見 backend/app/db/population_scope.py）。
#: ⚠ 理由是給複核的人看的——「忘了接母體」與「刻意全庫」在程式碼上長得一樣。
POPULATION_SCOPE_EXEMPT = {
    "list_zh_name_drafts":
        "公司治理跨 workspace：中文名草稿清單本來就是全庫，接母體反而看不到別包的待審",
    "count_company_normalization_queue":
        "同上：正規化候選池是全庫概念，不隸屬任一 workspace",
}

# 對照檔（xlsx/csv）表頭，2026-07-28 起與 DB 四欄一致（使用者：「對照檔也改四欄」）。
# ⚠ 順序即使用者指定的順序，不得調換；中文名與英文正式名分兩格填，
# 匯入不再以字元類別猜哪個是中文（混合字串會判錯且無人覆核）。
REQUIRED_COLUMNS = ("申請人代碼", "公司中文名稱", "正規化名稱", "別稱")

# CJK 範圍：只服務 `canonical` 舊單欄輸入的相容判斷（見 resolve_group_names）。
# 新流程由使用者在前端分兩格填中文／英文，不靠字元類別猜。
_CJK_RE = re.compile(r"[一-鿿]")


def notify_company_aliases_changed(cur: Any, *, action: str) -> None:
    """在目前 transaction 提交後通知瀏覽器重讀公司治理區。"""
    payload = {
        "kind": "data",
        "resource": "companyAliases",
        "action": action,
        "event_id": f"companyAliases:{uuid4().hex}",
    }
    cur.execute(
        "SELECT pg_notify('patent_events', %s)",
        (json.dumps(payload, separators=(",", ":")),),
    )


def normalize_lookup(value: str | None) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return " ".join(text.casefold().split())


# L2 疑似比對用的公司後綴（規格 applicant-code-grouping-spec.md 3-2）。
# 只剝**結尾**——中間出現的同字（如 "CO OP MACHINERY"）不動。
_LOOSE_SUFFIXES = (
    "co ltd", "co", "ltd", "limited", "inc", "incorporated", "llc", "lp",
    "corp", "corporation", "gmbh", "ag", "ab", "bv", "nv", "sa", "srl",
    "plc", "pte", "pty", "kk", "有限公司", "股份有限公司", "株式会社", "株式會社",
)
# 標點一律轉空白（不是刪除）——刪除會把 "A.B" 併成 "ab"，與 "AB" 撞在一起。
_LOOSE_PUNCT_RE = re.compile(r"[.,()\[\]{}'\"`;:/\\&+]")
# "doing business as"：⚠ **切斷**不是剝除，見 normalize_loose docstring。
_DBA_RE = re.compile(r"\bd\s*/?\s*b\s*/?\s*a\b|\bdba\b")


def normalize_loose(value: str | None) -> str | None:
    """L2 疑似比對鍵：在 `normalize_lookup` 之上去標點、剝結尾後綴、DBA 切斷。

    ⚠ **只供「疑似」提示，不得用於自動歸戶**（規格 3-3）：忽略後綴後
    「A CO., LTD.」與「A INC.」會得到同一個 key，但那可能是兩家不同法人。
    誤歸戶比漏歸戶難修——要人工找出來再拆開。

    ⚠ **DBA 是切斷不是剝除**（2026-07-30 實測資料驅動）：
    `SKI-ROW INC DBA ENERGYFIT` 的 DBA 在**中間**。當後綴剝掉會得到
    `SKI-ROW INC ENERGYFIT`——**不存在的名稱**，比不處理更糟。
    切斷後取前段 `SKI-ROW INC`（法人本名），再照常剝 INC。
    """
    text = normalize_lookup(value)
    if not text:
        return None
    # ① DBA 切斷：只取前段（法人本名）。
    cut = _DBA_RE.split(text, maxsplit=1)[0]
    # ② 標點轉空白後重新收斂。
    cleaned = " ".join(_LOOSE_PUNCT_RE.sub(" ", cut).split())
    # ③ 反覆剝結尾後綴（"X Co., Ltd." 去標點後是 "x co ltd"，要剝兩層）。
    changed = True
    while changed and cleaned:
        changed = False
        for suffix in _LOOSE_SUFFIXES:
            if cleaned == suffix:
                break
            if cleaned.endswith(" " + suffix):
                cleaned = cleaned[: -(len(suffix) + 1)].strip()
                changed = True
                break
    return cleaned or text


def load_alias_rows(path: Path, with_dropped: bool = False):
    """載入來源檔並正規化；with_dropped=True 時回傳 (rows, dropped)。"""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_alias_rows(path, with_dropped=with_dropped)
    if suffix == ".csv":
        return load_csv_alias_rows(path, with_dropped=with_dropped)
    raise ValueError(f"Unsupported company alias file format: {path.suffix}")


def load_csv_alias_rows(path: Path, with_dropped: bool = False):
    for encoding in ("utf-8-sig", "cp950", "big5", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return normalize_alias_rows(csv.DictReader(handle), with_dropped=with_dropped)
        except UnicodeError:
            continue
    raise UnicodeError(f"Cannot decode company alias CSV: {path}")


def load_xlsx_alias_rows(path: Path, with_dropped: bool = False):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    records = []
    for row in rows:
        record = {header: row[index] if index < len(row) else None for index, header in enumerate(headers) if header}
        records.append(record)
    return normalize_alias_rows(records, with_dropped=with_dropped)


def normalize_alias_rows(records: Any, with_dropped: bool = False):
    """讀來源檔（中文表頭）→ 輸出 SQL 具名參數用的英文 key，並依 DB 唯一索引去重。

    去重 key 必須與 DB 唯一索引 ux_company_aliases_code_lookup_confirmed 同一把
    ——即 (申請人代碼, normalize_lookup("別稱"))（後者為 lower + 壓空白）。
    2026-07-23 定案「代碼是公司收斂的依據」後唯一性下放到代碼層級：
    同一別稱字面可分屬不同代碼，只有「同代碼同別稱」才算重複。
    舊版用 (代碼, 舊單一名稱欄, 別稱) 三元組，抓不到只差大小寫／空白的變體，
    整批 executemany 會撞 UniqueViolation 導致全交易 rollback。

    同 key 多筆時保留**來源檔第一筆**：來源順序穩定可重現，不需臆測哪個字面
    「較正規」（大小寫偏好因公司而異，硬選反而是另一種寫死）。被丟棄的列
    一律回報，不靜默丟棄。
    """
    rows = []
    dropped = []
    seen: dict[tuple[str, str], dict[str, str]] = {}
    for record in records:
        company_code = clean_text(record.get("申請人代碼"))
        zh_name = clean_text(record.get("公司中文名稱"))
        normalized_name = clean_text(record.get("正規化名稱"))
        alias_name = clean_text(record.get("別稱"))
        # 兩個名稱欄各自可空，但**至少要有一個**——兩欄皆空的列沒有任何顯示名，
        # 寫進去只會讓顯示端 COALESCE 落空（0040 第②點刻意不加 CHECK 約束，
        # 由此處把關）。
        if not (zh_name or normalized_name) or not alias_name:
            continue
        key = (company_code or "", normalize_lookup(alias_name))
        if key in seen:
            kept = seen[key]
            dropped.append(
                {
                    "lookup_key": key[1],
                    "company_code": company_code,
                    # 警告訊息用的顯示名：中文優先、沒中文才英文（與顯示端同順位）。
                    "company_name": zh_name or normalized_name,
                    "alias_name": alias_name,
                    "kept_company_code": kept["company_code"],
                    "kept_alias_name": kept["alias_name"],
                }
            )
            continue
        row = {
            "company_code": company_code,
            "zh_name": zh_name or None,
            "normalized_name": normalized_name or None,
            "alias_name": alias_name,
        }
        seen[key] = row
        rows.append(row)
    if with_dropped:
        return rows, dropped
    return rows


def import_company_aliases(
    path: Path,
    dry_run: bool = False,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """匯入對照表；批內先依 DB 唯一索引同一把 key 去重，重複列記入 summary 警告。"""
    rows, dropped = load_alias_rows(path, with_dropped=True)
    summary = {
        "file": str(path),
        "file_format": path.suffix.lstrip(".").lower(),
        "rows": len(rows),
        "duplicate_dropped": len(dropped),
        "duplicate_warnings": dropped,
        "status": "dry_run" if dry_run else "pending",
    }
    if dry_run:
        return summary

    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for database import. Install psycopg[binary].") from exc

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        with conn.cursor() as cur:
            cur.executemany(
                """
                INSERT INTO derived_layer.company_aliases (
                    "申請人代碼", "公司中文名稱", "正規化名稱", "別稱", source_file
                )
                VALUES (
                    %(company_code)s, %(zh_name)s, %(normalized_name)s,
                    %(alias_name)s, %(source_file)s
                )
                -- 衝突目標＝ux_company_aliases_code_lookup_confirmed（同代碼同別稱）。
                -- 重跑同一檔時更新兩個名稱欄與來源，維持 idempotent。
                ON CONFLICT ("申請人代碼", alias_lookup_key) WHERE review_status = 'confirmed' DO UPDATE
                SET
                    "公司中文名稱" = EXCLUDED."公司中文名稱",
                    "正規化名稱" = EXCLUDED."正規化名稱",
                    "別稱" = EXCLUDED."別稱",
                    source_file = EXCLUDED.source_file,
                    imported_at = now()
                """,
                [{**row, "source_file": str(path)} for row in rows],
            )
        conn.commit()
    summary["status"] = "imported"
    return summary


# curation 裁決標記：company_aliases.source_file 帶此前綴＝該代碼已完成顯示名裁決
# （含「保留原文」的裁決），名稱治理管線不再把它列入待中文化建議。
CURATION_SOURCE_PREFIX = "display_name_curation"


# 名稱欄 → 對應代碼欄（2026-07-30 規格 applicant-code-grouping-spec.md 2-6）。
#
# ⚠ 原本自動歸戶只掃「申請人／標準化申請人」兩欄，而待補清單掃五欄——
# 專利權人／受讓人欄的名稱**看得見卻不會自動歸戶**（使用者：「範圍納入到
# 各種專利權人欄位都要」）。此表把兩者口徑統一。
#
# ⚠ `最近受讓人` 的代碼欄為 None：WIPS 匯出**沒有**受讓人代碼欄
# （mappings/wips.py 只有 `申請人代表碼` 與 `標準當前專利權人代碼` 兩個代碼欄）。
# 不得誤掛申請人代碼——那是不同欄位的代碼，掛了會把受讓人歸到申請人的組。
PEOPLE_NAME_CODE_COLUMNS: tuple[tuple[str, str | None], ...] = (
    ("申請人", "申請人代表碼"),
    ("標準化申請人", "申請人代表碼"),
    ("最近專利權人[US,JP,KR,CN,CA,AU]", "標準當前專利權人代碼[US,JP,KR,CN,CA,AU]"),
    ("標準當前專利權人[US,JP,KR,CN,CA,AU]", "標準當前專利權人代碼[US,JP,KR,CN,CA,AU]"),
    ("最近受讓人[US,KR,CN]", None),
)


#: 繁體欄名 → WIPS 原始（簡體）欄名。
#: 🔴 2026-08-03 實機驗收發現：`people` dict 保留 **WIPS 原始欄名**，
#: 而 `PEOPLE_NAME_CODE_COLUMNS` 寫死繁體——簡體檔一個欄位都對不上，
#: `build_people_pairs` 回傳 0 對，整條名稱治理管線**空轉且不報錯**
#: （匯入 summary 的 alias_variants 顯示 0，看起來像「沒有新變體」）。
#: ⚠ 反查表由 `wips.py` 的 `PEOPLE_FIELD_COLUMNS` **推導**，不另寫一份——
#: 本 bug 的成因正是同一組欄名在兩處各寫一次，其中一處只寫了繁體。
_TRADITIONAL_TO_SOURCE = {trad: simp for simp, trad in PEOPLE_FIELD_COLUMNS.items()}

#: 自動建組寫入的 source_type。⚠ 必須是 DB CHECK 白名單內的值，見上方 INSERT 註解。
AUTO_GROUP_SOURCE_TYPE = "wips_lookup"


def people_value(people: dict[str, Any], column: str | None) -> Any:
    """取 people 欄位值——繁體欄名與 WIPS 原始（簡體）欄名都認。

    ⚠ 兩種都要試：舊資料與 `alias_variant_sweep` 走 DB 讀出來的是繁體欄名，
    匯入路徑拿到的是 WIPS 原始欄名。只認一種就會有一條路徑靜默失效。
    """
    if column is None:
        return None
    if column in people:
        return people[column]
    source = _TRADITIONAL_TO_SOURCE.get(column)
    return people.get(source) if source else None


def build_people_pairs(people: dict[str, Any]) -> list[tuple[str | None, str]]:
    """從一列 patent_people 抽出 (代碼, 名稱) 配對，供名稱治理管線消費。

    匯入（wips_importer）與全量重掃（alias_variant_sweep）共用本函式，
    不各自維護一份掃描邏輯。

    ⚠ `A公司 | B公司` 多值要拆：WIPS 以 ` | ` 分隔同欄多個人／公司，
    不拆會把整串當成一家（實測庫內申請人 14 筆、專利權人 10 筆含此分隔符）。

    🔴 **只有每欄第一個名稱帶代碼，其餘一律 None**（2026-07-30 使用者定案）：
    WIPS 的代碼欄是**整列一個**，拆名稱時無從得知哪個名稱對應那個代碼。
    若每筆都掛同一代碼，第二個名稱（常是自然人，如 `A公司 | Zeng Qing` 的
    `Zeng Qing`）會被**自動併進公司組**——那跨過「系統不預先分組」的紅線，
    即使結果可能正確也該由使用者按一下。
    依據：WIPS 慣例第一個是主申請人（`refresh_report_patent_base` 的
    `split_part(..., '|', 1)` 同一假設）。其餘名稱走無代碼路徑進待補清單。

    ⚠ 個人名（自然人）不在此過濾——它們照常進待補清單，由使用者決定歸入或忽略。
    """
    pairs: list[tuple[str | None, str]] = []
    seen: set[tuple[str | None, str]] = set()
    for name_col, code_col in PEOPLE_NAME_CODE_COLUMNS:
        raw_name = clean_text(people_value(people, name_col))
        if not raw_name:
            continue
        col_code = clean_text(people_value(people, code_col)) if code_col else None
        # 每欄各自算「第一個」——專利權人欄的第一個要帶它自己的代碼欄。
        is_first = True
        for part in raw_name.split("|"):
            name = clean_text(part)
            if not name:
                continue
            code = col_code if is_first else None
            is_first = False
            key = (code, name)
            if key in seen:
                continue
            seen.add(key)
            pairs.append(key)
    return pairs


def govern_company_names(
    pairs: list[tuple[str | None, str | None]],
    source_label: str = "variant_intake",
    connect_kwargs: dict[str, Any] | None = None,
    standardized_names: dict[str, str] | None = None,
) -> dict[str, Any]:
    """名稱治理管線核心：變體註冊＋待中文化偵測，匯入與 sweep 共用同一套邏輯。

    輸入 (code, name) 集合（2026-07-21 定案「單一名稱治理管線」）：
    - code 在 company_aliases 對到唯一名稱組（中文名＋英文正式名）→ 新變體補一列別稱，
      兩個名稱欄沿用既有值；
      既有別稱（normalize 後相同）跳過。
    - code 不在表中 → **自動建待確認組**（2026-07-30 規格 applicant-code-grouping-spec.md
      批次 b）：WIPS 同時給代碼與 `標準化申請人`，四欄中三欄可直接推導，只有中文名
      （市場慣用名＝判斷不是資料）留空。標 `review_status='review_required'`，
      既有消費端（待補清單、報表顯示）都只吃 `confirmed`，故不污染正式資料，
      且該名稱**仍留在待補清單**直到使用者確認。
      ⚠ 不用 `ai_suggested`——本路徑是確定性規則，無 AI 參與（使用者明示）。
    - code 對到多組名稱（conflicting_code）→ 進 manual_review，不自行合併、不寫表。
      ⚠ 衝突代碼**不自動建組**：資料本身有問題，猜哪一組都可能錯。

    standardized_names＝{代碼: WIPS 標準化申請人}，供建組時填英文正式名；
    未提供或該代碼缺值時退回該筆的原始名稱（不因缺一欄就整組不建）。
    - needs_zh_name：本批涉及且已在對照表的 code 中，canonical 顯示名不含 CJK 字元
      （尚無市場慣用中文名）且未經 curation 裁決者——該代碼不存在
      source_file LIKE 'display_name_curation%' 的列；「保留原文」也是正式裁決，不重複浮現。
    - 只寫 derived_layer.company_aliases；不碰 raw/core 原始專利值。
    """
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for database import. Install psycopg[binary].") from exc

    from backend.app.db.connection import get_connection_kwargs

    inserted = 0
    skipped = 0
    manual: list[dict[str, str]] = []
    needs_zh: list[dict[str, str]] = []
    # 本批自動建立的待確認組（規格批次 b）；供匯入 summary 顯示「待確認代碼組 N 組」。
    created_groups: list[dict[str, str]] = []
    # L2 疑似同一家（規格批次 a）：只回報供使用者確認，**不寫入對照表**。
    suspected: list[dict[str, str]] = []
    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        # 2026-07-28 四欄定案：只讀 `公司中文名稱`／`正規化名稱`
        # （拆欄前的單一名稱欄已於 0041 移除，不再有 fallback）。
        rows = conn.execute(
            'SELECT "申請人代碼", "公司中文名稱", "正規化名稱", "別稱" '
            "FROM derived_layer.company_aliases "
            'WHERE "申請人代碼" IS NOT NULL'
        ).fetchall()
        # code → 既有 (中文名, 英文正式名) 集合／normalize 後別稱集合
        # ⚠ 以「(中文, 英文) 這一組」為單位判重，不是各欄獨立比對——只比其中一欄
        # 會把「中文相同、英文不同」的兩組當成同一組而靜默合併。
        names_by_code: dict[str, set[tuple[str, str]]] = {}
        aliases_by_code: dict[str, set[str]] = {}
        for code, zh_name, normalized, alias in rows:
            zh = clean_text(zh_name) or ""
            en = clean_text(normalized) or ""
            if zh or en:
                names_by_code.setdefault(code, set()).add((zh, en))
            else:
                # 三欄全空的殘列不參與 canonical 判定，但也不讓該代碼變成
                # unknown_code——先登記空集合，由下方 len 檢查處理。
                names_by_code.setdefault(code, set())
            norm = normalize_lookup(alias)
            if norm:
                aliases_by_code.setdefault(code, set()).add(norm)

        # 無代碼時的名稱索引（規格批次 a）：L1 精確鍵與 L2 寬鬆鍵各一份。
        # ⚠ L2 只用於「疑似」提示，不自動寫入（見 normalize_loose docstring）。
        code_by_exact: dict[str, str] = {}
        code_by_loose: dict[str, str] = {}
        for code_key, alias_set in aliases_by_code.items():
            for norm in alias_set:
                code_by_exact.setdefault(norm, code_key)
        for code_key, name_set in names_by_code.items():
            for zh, en in name_set:
                for official in (zh, en):
                    norm = normalize_lookup(official)
                    if norm:
                        code_by_exact.setdefault(norm, code_key)
        for norm, code_key in code_by_exact.items():
            loose = normalize_loose(norm)
            if loose:
                code_by_loose.setdefault(loose, code_key)

        batch_codes: set[str] = set()
        for raw_code, raw_variant in pairs:
            code = clean_text(raw_code)
            variant = clean_text(raw_variant)
            if not variant:
                continue
            if not code:
                # 🆕 無代碼 → 名稱兩級比對（規格批次 a，2026-07-30）。
                # 實測 60 筆專利中 57 筆（95%）無代碼，使用者手建的 20 組 TEMP
                # 原本對新資料**完全沒有作用**——只認代碼的話那些組形同虛設。
                norm_variant = normalize_lookup(variant)
                if not norm_variant:
                    continue
                hit = code_by_exact.get(norm_variant)
                if hit:
                    # L1 完全命中 → 歸入該組。已有這個別稱就跳過，否則補一列。
                    if norm_variant in aliases_by_code.get(hit, set()):
                        skipped += 1
                        continue
                    names = names_by_code.get(hit) or set()
                    canonical_zh, canonical_en = next(iter(names)) if len(names) == 1 else ("", "")
                    conn.execute(
                        'INSERT INTO derived_layer.company_aliases '
                        '("申請人代碼", "公司中文名稱", "正規化名稱", "別稱", '
                        " source_file, source_type) "
                        # ⚠ 明確指定 source_type，不吃 DB 預設值 `excel_seed`
                        # ——這條路的來源是 **WIPS 匯入**，不是 excel 對照檔。
                        # 2026-08-03 實機驗收看到別稱列標成 excel_seed 才發現。
                        f"VALUES (%s, %s, %s, %s, %s, '{AUTO_GROUP_SOURCE_TYPE}') "
                        'ON CONFLICT ("申請人代碼", alias_lookup_key) '
                        "WHERE review_status = 'confirmed' DO NOTHING",
                        (hit, canonical_zh or None, canonical_en or None, variant, source_label),
                    )
                    aliases_by_code.setdefault(hit, set()).add(norm_variant)
                    code_by_exact.setdefault(norm_variant, hit)
                    inserted += 1
                    continue
                loose_variant = normalize_loose(variant)
                loose_hit = code_by_loose.get(loose_variant) if loose_variant else None
                if loose_hit:
                    # ⚠ L2 疑似**只提示不寫入**：忽略後綴後「A CO., LTD.」與
                    # 「A INC.」同 key，可能是兩家不同法人。誤歸戶比漏歸戶難修。
                    suspected.append({
                        "company_code": loose_hit,
                        "alias_name": variant,
                        "reason": "suspected_same_company",
                    })
                # L3 都不命中：不寫入、不回報——維持既有行為（進待補清單）。
                continue
            batch_codes.add(code)
            names = names_by_code.get(code)
            if names is None:
                # 🆕 未建組的真代碼 → 自動建待確認組（規格批次 b）。
                # 英文正式名優先取 WIPS 標準化申請人；缺就退回原始寫法
                # （⚠ 不因缺一欄就整組不建，那等於回到「只丟 manual」的現況）。
                canonical_en = clean_text((standardized_names or {}).get(code)) or variant
                conn.execute(
                    'INSERT INTO derived_layer.company_aliases '
                    '("申請人代碼", "公司中文名稱", "正規化名稱", "別稱", '
                    " source_file, source_type, review_status) "
                    # 🔴 source_type 必須是 CHECK 白名單內的值
                    # （excel_seed／wips_lookup／manual／ai_suggested）。
                    # 2026-08-03 實機驗收：原本寫 'import'，**不在白名單**，
                    # 一送出就 CheckViolation——規格 2-2 節誤記為「既有值」。
                    # ⚠ 用 `wips_lookup`：本路徑就是「拿 WIPS 給的代碼與標準化名建組」。
                    # ⚠ 不可用 `ai_suggested`——本線是確定性規則、無 AI 參與（使用者明示）；
                    #   也不是 `manual`（不是人工建的）、不是 `excel_seed`（那是初始種子檔）。
                    f"VALUES (%s, %s, %s, %s, %s, '{AUTO_GROUP_SOURCE_TYPE}', 'review_required') "
                    'ON CONFLICT ("申請人代碼", alias_lookup_key) '
                    "WHERE review_status = 'confirmed' DO NOTHING",
                    # ⚠ 中文名一律 None：市場慣用名是判斷不是資料，不得自動填。
                    (code, None, canonical_en, variant, f"auto:{source_label}"),
                )
                # 同批同代碼的第二個名稱要能續掛，故就地登記為已知組。
                names_by_code[code] = {("", canonical_en)}
                aliases_by_code.setdefault(code, set()).add(normalize_lookup(variant))
                created_groups.append({
                    "company_code": code,
                    "normalized_name": canonical_en,
                    "alias_name": variant,
                })
                continue
            if len(names) > 1:
                manual.append({"company_code": code, "alias_name": variant, "reason": "conflicting_code"})
                continue
            if not names:
                # 該代碼在表中有列，但兩個名稱欄都是空的（殘列）——沒有可沿用的
                # canonical，寫進去會是「有代碼有別稱、沒有任何顯示名」的新殘列。
                # 丟人工處理，不自行編名。⚠ 少了這條 `next(iter(names))` 會 StopIteration
                # 直接中斷整批，而非只跳過這一筆。
                manual.append({"company_code": code, "alias_name": variant, "reason": "no_canonical_name"})
                continue
            norm_variant = normalize_lookup(variant)
            if norm_variant in aliases_by_code.get(code, set()):
                skipped += 1
                continue
            canonical_zh, canonical_en = next(iter(names))  # 唯一既有名稱組，直接沿用
            conn.execute(
                # ⚠ 衝突目標改為 partial unique index（2026-07-28，0040 拆四欄）：
                # 舊三元組 UNIQUE (代碼, 舊名稱欄, 別稱) 已隨拆欄移除，這裡是全 repo
                # 唯一依賴它的寫入路徑。改用與 import_company_aliases／
                # apply_confirmed_display_names 同一把 key，唯一性語意不變。
                # ⚠ 名稱欄一併改寫四欄口徑：中文進 `公司中文名稱`、英文正式名進
                # `正規化名稱`（0041 起表上只有這四欄）。
                'INSERT INTO derived_layer.company_aliases '
                '("申請人代碼", "公司中文名稱", "正規化名稱", "別稱", '
                " source_file, source_type) "
                # ⚠ 同上：已知代碼補別稱同樣來自 WIPS 匯入，不是 excel 對照檔。
                f"VALUES (%s, %s, %s, %s, %s, '{AUTO_GROUP_SOURCE_TYPE}') "
                'ON CONFLICT ("申請人代碼", alias_lookup_key) '
                "WHERE review_status = 'confirmed' DO NOTHING",
                (code, canonical_zh or None, canonical_en or None, variant, source_label),
            )
            aliases_by_code.setdefault(code, set()).add(norm_variant)
            inserted += 1

        # 待中文化偵測：只看本批涉及的 code；尚無中文名且該代碼沒有任何 curation
        # 裁決列（source_file LIKE 'display_name_curation%'）才浮現。
        # ⚠ 2026-07-28 四欄定案：判準由「舊單一名稱欄不含 CJK」改為
        # 「`公司中文名稱` 為空」。舊判準在拆欄後會失效——新列的舊名稱欄是
        # NULL，而 `NULL !~ '...'` 在 PostgreSQL 得到 NULL（非 TRUE），
        # WHERE 不成立 → 該代碼**永遠不會浮現待中文化**（靜默失效）。
        # 待中文化的列必然沒有中文名，故顯示名直接取英文正式名。
        if batch_codes:
            zh_rows = conn.execute(
                'SELECT DISTINCT ca."申請人代碼", '
                '       COALESCE(NULLIF(BTRIM(ca."正規化名稱"), \'\'), \'\') '
                "FROM derived_layer.company_aliases ca "
                'WHERE ca."申請人代碼" = ANY(%s) '
                '  AND NULLIF(BTRIM(COALESCE(ca."公司中文名稱", \'\')), \'\') IS NULL '
                "  AND NOT EXISTS ("
                "      SELECT 1 FROM derived_layer.company_aliases d "
                '      WHERE d."申請人代碼" = ca."申請人代碼" AND d.source_file LIKE %s'
                "  ) "
                "ORDER BY 1, 2",
                (sorted(batch_codes), f"{CURATION_SOURCE_PREFIX}%"),
            ).fetchall()
            needs_zh = [{"company_code": c, "company_name": n} for c, n in zh_rows]
        conn.commit()

    return {
        "inserted": inserted,
        "skipped_existing": skipped,
        "manual_review": manual,
        "needs_zh_name": needs_zh,
        "created_groups": created_groups,
        "suspected": suspected,
    }


def register_known_code_variants(
    pairs: list[tuple[str | None, str | None]],
    source_label: str = "variant_intake",
    connect_kwargs: dict[str, Any] | None = None,
    standardized_names: dict[str, str] | None = None,
) -> dict[str, Any]:
    """WIPS code 的名稱變體補入唯一對照表（薄包裝，委派名稱治理管線）。

    保留原簽章供 wips_importer 匯入時呼叫；實際邏輯統一在 govern_company_names，
    匯入摘要因此自動帶出 needs_zh_name（新公司進庫即浮現待中文化）。

    ⚠ 新增參數**必須逐層轉傳**：本專案已兩次踩「中間層漏接參數」的靜默失敗
    （前端送 `aliases` 後端欄位是 `variants`、`report_keys` 被 Pydantic 丟棄），
    兩次都是頭尾對、中間斷，測試照樣全綠。
    """
    return govern_company_names(
        pairs,
        source_label=source_label,
        connect_kwargs=connect_kwargs,
        standardized_names=standardized_names,
    )


def resolve_group_names(spec: dict[str, Any]) -> tuple[str | None, str | None]:
    """從一組 mapping spec 取出 (中文正式名, 英文正式名)。

    2026-07-28 四欄拆分後 mapping 的名稱欄有兩個鍵：
    - `zh_name`         → `公司中文名稱`（中文正式名）
    - `normalized_name` → `正規化名稱`（英文正式名）

    ⚠ 相容舊鍵 `canonical`：既有呼叫端（中文名確認端點 confirm_drafts）只給一個
    顯示名，語意上「有 CJK 就是中文名、否則是英文正式名」。這裡的字元類別判斷
    **只用於相容舊單欄輸入**，不是新流程的判斷依據——新流程由使用者在前端分兩格
    填，不靠猜（使用者第③點否決的正是「自動判斷含 CJK 就是中文名」的資料遷移）。
    兩欄都可空（使用者第②點），全空時回 (None, None)，由呼叫端決定要不要寫。
    """
    zh = clean_text(spec.get("zh_name"))
    en = clean_text(spec.get("normalized_name"))
    if zh or en:
        return zh, en
    legacy = clean_text(spec.get("canonical"))
    if not legacy:
        return None, None
    return (legacy, None) if _CJK_RE.search(legacy) else (None, legacy)


def apply_confirmed_display_names(
    mapping: dict[str, dict[str, Any]],
    source_label: str,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """套用使用者已裁決的公司顯示名（curation 落地機制；載體＝DB 本身，不經 CSV）。

    mapping = {申請人代碼: {"zh_name": 中文正式名, "normalized_name": 英文正式名,
                            "aliases": [變體, ...]}}
    （2026-07-28 四欄拆分；舊形狀 `{"canonical": 顯示名, ...}` 仍相容，
      見 resolve_group_names）

    規則（2026-07-21 公司顯示名原則定案；2026-07-23 隨代碼收斂調整；2026-07-28 拆四欄）：
    - **兩個正式名自身也納入別稱**，確保使用者填的中文名／英文名字面在專利表可精確
      命中（原本只納入單一 canonical；拆欄後兩個都要，否則另一個字面命不中）。
    - 批內按 (代碼, normalize_lookup(別稱)) 去重，與 DB 唯一索引
      ux_company_aliases_code_lookup_confirmed 同一把 key。去重**不跨 code**
      ——跨 code 去重會在同一批內誤丟別家公司的列。
      🔴 2026-08-18 更正：本段原寫「不同代碼可有同一別稱字面」，那已**不再成立**。
      migration 0052 加了 `ux_company_aliases_lookup_single_code`
      （partial unique on alias_lookup_key WHERE confirmed）——一個別稱只能屬於
      一個代碼，否則歸戶取決於查詢順序（實例：TTI Macao 同時掛在 UN164421 與
      UN240278 下，WIPS 確認只屬後者）。
      ⚠ 因此本函式寫入撞名時會拋 UniqueViolation；呼叫端負責翻成 409。
      ⚠ 「一家公司多個代碼」仍完全合法（創科集團有 4 個），由 company_groups 收攏
      ——被禁止的是「一個別稱多個代碼」，兩者不同。
    - (代碼, lookup key) 已存在 → UPDATE 該列 re-canonicalize：改掛新的中文／英文
      正式名與 source_file，review_status 一律轉 'confirmed'；不插新列，不撞唯一索引。
      查詢必須同時比對代碼，否則在「一別稱多公司」下會取到別家公司的列。
    - 不存在 → INSERT 一列 confirmed（source_type='manual'＝人工裁決）。
    - 兩個正式名皆空的組直接略過（使用者第②點：可空、不擋、不報錯；但沒有任何名字
      也就沒有東西可寫）。
    - 只寫 derived_layer.company_aliases；原始專利表不碰。
    回傳 {"inserted", "updated", "dedup_dropped"}。

    ⚠ 這是 company_aliases 唯一的 confirmed 寫入點。去重、re-canonicalize、
    review_status 轉換的規則只有這一份，其他模組一律委派，不得自寫 SQL。
    """
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for database import. Install psycopg[binary].") from exc

    from backend.app.db.connection import get_connection_kwargs

    # 批內去重：唯一索引已含代碼，去重 key 同步改為 (code, lookup key)，
    # 不同代碼的同一別稱字面各自保留，不再互相排擠。
    entries: list[tuple[str, str | None, str | None, str]] = []  # (code, 中文名, 英文名, 別稱)
    seen_keys: set[tuple[str, str]] = set()
    dedup_dropped = 0
    for raw_code, spec in mapping.items():
        code = clean_text(raw_code)
        zh_name, en_name = resolve_group_names(spec)
        if not code or not (zh_name or en_name):
            continue
        # 兩個正式名都納入別稱（各自的字面都要能命中），再接使用者填的變體。
        for raw_alias in [zh_name, en_name, *spec.get("aliases", [])]:
            alias = clean_text(raw_alias)
            if not alias:
                continue
            key = (code, normalize_lookup(alias))
            if key in seen_keys:
                dedup_dropped += 1
                continue
            seen_keys.add(key)
            entries.append((code, zh_name, en_name, alias))

    inserted = 0
    updated = 0
    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        for code, zh_name, en_name, alias in entries:
            # 以 DB 端與 generated 欄完全相同的正規化運算式找既有列。
            # 必須同時比對「申請人代碼」：唯一索引已是 (代碼, lookup key)，
            # 只用別稱會在「一別稱多公司」時取到別家公司的列並把它改掛到本代碼。
            # 多列同 key（review_required 重複）時優先更新 confirmed 列，其餘不動。
            row = conn.execute(
                "SELECT id FROM derived_layer.company_aliases "
                r"WHERE alias_lookup_key = lower(regexp_replace(btrim(%s), '\s+', ' ', 'g')) "
                'AND "申請人代碼" IS NOT DISTINCT FROM %s '
                "ORDER BY (review_status = 'confirmed') DESC, id LIMIT 1",
                (alias, code),
            ).fetchone()
            if row:
                conn.execute(
                    'UPDATE derived_layer.company_aliases SET "公司中文名稱" = %s, '
                    '"正規化名稱" = %s, '
                    "source_file = %s, review_status = 'confirmed', source_type = 'manual', "
                    "updated_at = now() WHERE id = %s",
                    (zh_name, en_name, source_label, row[0]),
                )
                updated += 1
            else:
                conn.execute(
                    'INSERT INTO derived_layer.company_aliases ("申請人代碼", "公司中文名稱", '
                    '"正規化名稱", "別稱", source_file, source_type, review_status) '
                    "VALUES (%s, %s, %s, %s, %s, 'manual', 'confirmed')",
                    (code, zh_name, en_name, alias, source_label),
                )
                inserted += 1
        conn.commit()

    return {"inserted": inserted, "updated": updated, "dedup_dropped": dedup_dropped}


# AI 中文名草稿清單（三態的「草稿待確認」）：只列 ai_suggested 列。
# 一併帶出：
#   - draft_name＝AI 產的中文名草稿（2026-07-28 起讀 `公司中文名稱`；keep_original
#     裁決時該欄為 NULL，前端據 verdict 顯示「查無，保留原文」）
#   - source_name＝AI 的輸入英文名（`正規化名稱`），供對照
#   - verdict＝translated／keep_original，供前端區分「有中文名」與「查無保留原文」
#   - original_name＝收斂前的原始字面（該代碼的別稱），確認中文名時的對照依據
# ⚠ 不帶「該代碼現行顯示名」：它與專利表「申請人」欄同源（皆走 code_alias_names →
# COALESCE 第一順位），而確認介面就掛在瀏覽專利頁、表格本身即顯示該欄，重複無益。
# 以代碼為單位，一代碼至多一草稿列（write_drafts 先刪後插保證）。
_LIST_DRAFTS_SQL = """
    SELECT d."申請人代碼" AS code,
           d."公司中文名稱" AS draft_name,
           d."正規化名稱" AS source_name,
           d.wips_metadata_json->>'zh_name_verdict' AS verdict,
           -- ⚠ 表沒有 created_at（實機 HTTP 500：UndefinedColumn）。時間欄是
           -- imported_at（匯入）與 updated_at（更新）；草稿為「一代碼至多一列、
           -- 重跑 UPDATE 收斂」，故顯示「何時產的」取 updated_at。
           d.updated_at AS created_at,
           (
               SELECT mode() WITHIN GROUP (ORDER BY c."別稱")
               FROM derived_layer.company_aliases c
               WHERE c."申請人代碼" = d."申請人代碼"
                 AND c.review_status = 'confirmed'
                 AND NULLIF(BTRIM(c."別稱"), '') IS NOT NULL
           ) AS original_name
    FROM derived_layer.company_aliases d
    WHERE d.review_status = 'ai_suggested'
      AND NULLIF(BTRIM(d."申請人代碼"), '') IS NOT NULL
    ORDER BY d."申請人代碼"
    LIMIT %(limit)s OFFSET %(offset)s
"""

# 總筆數：分頁時前端要知道還有多少沒看（與清單同一 WHERE，不另立條件）。
_COUNT_DRAFTS_SQL = """
    SELECT count(*) AS total
    FROM derived_layer.company_aliases d
    WHERE d.review_status = 'ai_suggested'
      AND NULLIF(BTRIM(d."申請人代碼"), '') IS NOT NULL
"""


def list_zh_name_drafts(
    *,
    limit: int = 100,
    offset: int = 0,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """列出待確認的 AI 中文名草稿（三態的「草稿待確認」）。

    回 {"items": [{code, draft_name, verdict, created_at, original_name}, ...], "total": N}。
    verdict＝translated（AI 找到市場慣用中文名）／keep_original（查無，保留英文原文）。
    original_name＝該代碼收斂前的原始字面，供使用者判斷這個中文名對不對。

    分頁（規格 B）：代碼數量可觀時不一次全吐；total 走 count(*) 而非 len(items)，
    否則第二頁之後前端算不出還剩幾筆。清單與計數同一 WHERE 條件，不會對不上。
    """
    import psycopg
    from psycopg.rows import dict_row

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        with conn.cursor(row_factory=dict_row) as cur:
            cur.execute(_LIST_DRAFTS_SQL, {"limit": limit, "offset": offset})
            items = cur.fetchall()
            cur.execute(_COUNT_DRAFTS_SQL)
            total = int(cur.fetchone()["total"])
    return {"items": items, "total": total}


def get_draft_names(
    code: str, connect_kwargs: dict[str, Any] | None = None
) -> tuple[str | None, str | None]:
    """取某代碼草稿列的 (中文名, 英文正式名)。

    前端只送 code＋action，草稿名由後端自己查——避免前端竄改或送到過期的草稿名。

    ⚠ 2026-07-28 拆四欄後回**兩個值**：`keep_original` 草稿的中文欄是 NULL，
    此時仍要能確認（英文正式名照樣寫進去、顯示退英文），故不能像舊版那樣
    「查不到名字就 422」。
    """
    import psycopg

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        row = conn.execute(
            'SELECT "公司中文名稱", "正規化名稱" FROM derived_layer.company_aliases '
            "WHERE \"申請人代碼\" = %s AND review_status = 'ai_suggested' LIMIT 1",
            (code,),
        ).fetchone()
    if not row:
        return None, None
    return clean_text(row[0]), clean_text(row[1])


_COMPANY_NORMALIZATION_CANDIDATES_SQL = r"""
    WITH raw_names AS (
        SELECT x.raw_name, x.source_field, pp.patent_id
        FROM core_layer.patent_people pp
        CROSS JOIN LATERAL (VALUES
            (NULLIF(BTRIM(pp."申請人"), ''), '申請人'),
            (NULLIF(BTRIM(pp."標準化申請人"), ''), '標準化申請人'),
            (NULLIF(BTRIM(pp."最近專利權人[US,JP,KR,CN,CA,AU]"), ''), '最近專利權人'),
            (NULLIF(BTRIM(pp."標準當前專利權人[US,JP,KR,CN,CA,AU]"), ''), '標準當前專利權人'),
            (NULLIF(BTRIM(pp."最近受讓人[US,KR,CN]"), ''), '最近受讓人')
        ) AS x(raw_name, source_field)
        WHERE x.raw_name IS NOT NULL
    ),
    names AS (
        SELECT lower(regexp_replace(BTRIM(part), '\s+', ' ', 'g')) AS lookup_key,
               BTRIM(part) AS raw_name,
               r.source_field,
               r.patent_id
        FROM raw_names r
        CROSS JOIN LATERAL regexp_split_to_table(r.raw_name, '\s*\|\s*') AS part
        WHERE NULLIF(BTRIM(part), '') IS NOT NULL
    )
    SELECT md5(n.lookup_key) AS ref_hash,
           n.lookup_key,
           min(n.raw_name) AS raw_name,
           array_agg(DISTINCT n.source_field ORDER BY n.source_field) AS source_fields,
           count(DISTINCT n.patent_id) AS patent_count,
           a.last_asked_at
    FROM names n
    -- 排隊（2026-08-18）：候選是即時算出來的，沒有實體列可蓋章，所以查不到證據的
    -- 候選每跑一次就再燒一次。蓋章表補上這個事實。
    LEFT JOIN derived_layer.company_normalization_asked a
           ON a.lookup_key = n.lookup_key
    WHERE NOT EXISTS (
        SELECT 1 FROM derived_layer.company_aliases ca
        WHERE ca.review_status = 'confirmed'
          AND lower(regexp_replace(BTRIM(ca."別稱"), '\s+', ' ', 'g')) = n.lookup_key
    )
      AND NOT EXISTS (
        SELECT 1 FROM derived_layer.company_aliases ca
        WHERE ca.review_status = 'ai_suggested'
          AND lower(regexp_replace(BTRIM(ca."別稱"), '\s+', ' ', 'g')) = n.lookup_key
          AND ca.source_file = 'ai:company_normalization_suggestion'
    )
    GROUP BY n.lookup_key, a.last_asked_at, a.asked_patent_count
    -- 資格（使用者裁決「乙」）：問過的只有在該名稱又有新專利進來時才重問。
    -- ⚠ 少了這一段就只是「延後重問」——輪完一圈後那批結構性查不到的
    --   （多為自然人）會再燒一次。
    HAVING a.asked_patent_count IS NULL
        OR count(DISTINCT n.patent_id) > a.asked_patent_count
    -- 順序：沒問過的一律排前面。⚠ NULLS FIRST 這一個子句就是
    --   「全部輪過一遍才會有人被問第二次」的全部實作，不需要輪次計數器。
    ORDER BY a.last_asked_at ASC NULLS FIRST,
             count(DISTINCT n.patent_id) DESC,
             min(n.raw_name)
    LIMIT %(limit)s
"""

_COMPANY_NORMALIZATION_TARGETS_SQL = """
    SELECT md5("申請人代碼") AS ref_hash,
           "申請人代碼" AS code,
           max(NULLIF(BTRIM("公司中文名稱"), '')) AS zh_name,
           max(NULLIF(BTRIM("正規化名稱"), '')) AS normalized_name,
           count(*) AS alias_count
    FROM derived_layer.company_aliases
    WHERE review_status = 'confirmed'
      AND NULLIF(BTRIM("申請人代碼"), '') IS NOT NULL
    GROUP BY "申請人代碼"
    HAVING max(NULLIF(BTRIM("公司中文名稱"), '')) IS NOT NULL
        OR max(NULLIF(BTRIM("正規化名稱"), '')) IS NOT NULL
    ORDER BY COALESCE(max(NULLIF(BTRIM("公司中文名稱"), '')),
                      max(NULLIF(BTRIM("正規化名稱"), '')),
                      "申請人代碼")
"""

_LIST_COMPANY_NORMALIZATION_SUGGESTIONS_SQL = """
    SELECT id,
           "申請人代碼" AS company_code,
           "公司中文名稱" AS suggested_zh_name,
           "正規化名稱" AS suggested_normalized_name,
           "別稱" AS raw_name,
           wips_metadata_json AS metadata,
           updated_at
    FROM derived_layer.company_aliases
    WHERE review_status = 'ai_suggested'
      AND source_file = 'ai:company_normalization_suggestion'
    ORDER BY updated_at DESC, id
    LIMIT %(limit)s OFFSET %(offset)s
"""

_COUNT_COMPANY_NORMALIZATION_SUGGESTIONS_SQL = """
    SELECT count(*) AS total
    FROM derived_layer.company_aliases
    WHERE review_status = 'ai_suggested'
      AND source_file = 'ai:company_normalization_suggestion'
"""


def list_company_normalization_candidates(
    *,
    limit: int | None = 100,
    connect_kwargs: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """列出 AI 可查證的未歸戶原始變體；ref 為 opaque，不含可寫 WIPS code。"""
    import psycopg
    from psycopg.rows import dict_row

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs()), row_factory=dict_row) as conn:
        rows = conn.execute(
            _COMPANY_NORMALIZATION_CANDIDATES_SQL,
            {"limit": int(limit or 100)},
        ).fetchall()
    return [
        {
            "candidate_ref": f"cand:{row['ref_hash'][:16]}",
            # ⚠ 內部鍵：蓋章時原樣寫回，不在 Python 重算正規化運算式。
            #   送 CLI 前必須被 `_public_candidates` 投影掉（見 runner）。
            "lookup_key": row["lookup_key"],
            "raw_name": row["raw_name"],
            "candidate_type": "company_or_person",
            "source_fields": row.get("source_fields") or [],
            "patent_count": int(row.get("patent_count") or 0),
        }
        for row in rows
    ]


def mark_company_normalization_asked(
    entries: list[dict[str, Any]],
    *,
    run_id: int | None = None,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, int]:
    """蓋章：記下這批候選被問過、當時件數與結果。

    ⚠ 只有**完成的段**才呼叫這裡。契約錯誤代表協定壞了，不是「這些候選查不到
    證據」；蓋了章會把程式問題當成資料結論，而件數不變就再也回不到隊列——
    一批候選因為一次程式錯誤永久消失，且沒有任何訊息。
    """
    if not entries:
        return {"stamped": 0}
    import psycopg

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        for entry in entries:
            conn.execute(
                """
                INSERT INTO derived_layer.company_normalization_asked
                    (lookup_key, last_asked_at, last_run_id,
                     asked_patent_count, outcome)
                VALUES (%s, now(), %s, %s, %s)
                ON CONFLICT (lookup_key) DO UPDATE SET
                    last_asked_at = now(),
                    last_run_id = EXCLUDED.last_run_id,
                    asked_patent_count = EXCLUDED.asked_patent_count,
                    outcome = EXCLUDED.outcome
                """,
                (entry["lookup_key"], run_id,
                 int(entry.get("patent_count") or 0),
                 entry.get("outcome") or "no_evidence"),
            )
        conn.commit()
    return {"stamped": len(entries)}


def count_company_normalization_queue(
    *,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, int]:
    """尚未查證的候選數量，分「從未查證」與「有新專利待重查」。

    ⚠ 分批若不揭露剩餘量，使用者會把「這批做完」讀成「全部做完」——
    與 2026-08-18 修掉的跳過靜默是同一類缺席型偏差。
    """
    import psycopg
    from psycopg.rows import dict_row

    from backend.app.db.connection import get_connection_kwargs

    # ⚠ 在 SQL 裡聚合，不要把整批候選撈回來再數——查詢本身約 450ms，
    #   多傳幾百列只是白花。候選數會隨新資料成長，這裡不能用「先全撈再算」。
    sql = (f"SELECT count(*) AS remaining, "
           f"count(*) FILTER (WHERE last_asked_at IS NULL) AS never_asked "
           f"FROM ({_COMPANY_NORMALIZATION_CANDIDATES_SQL}) AS q")
    with psycopg.connect(**(connect_kwargs or get_connection_kwargs()),
                         row_factory=dict_row) as conn:
        row = conn.execute(sql, {"limit": None}).fetchone()
    remaining = int(row["remaining"] or 0)
    never = int(row["never_asked"] or 0)
    return {"remaining": remaining, "never_asked": never,
            "recheck": remaining - never}


def list_company_normalization_targets(
    *,
    connect_kwargs: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    """列出 Backend 私有 target 白名單；CLI 只會看到 target_ref 與名稱，不看 code。"""
    import psycopg
    from psycopg.rows import dict_row

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs()), row_factory=dict_row) as conn:
        rows = conn.execute(_COMPANY_NORMALIZATION_TARGETS_SQL).fetchall()
    return [
        {
            "target_ref": f"target:{row['ref_hash'][:16]}",
            "code": row["code"],
            "zh_name": row.get("zh_name"),
            "normalized_name": row.get("normalized_name"),
            "alias_count": int(row.get("alias_count") or 0),
        }
        for row in rows
    ]


def ingest_company_normalization_suggestions(
    suggestions: list[dict[str, Any]],
    *,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, int]:
    """把 AI 驗證後的建議寫成 `ai_suggested`；不改 confirmed、不碰 raw/core。"""
    if not suggestions:
        return {"inserted": 0}
    import psycopg
    from psycopg.types.json import Jsonb

    from backend.app.db.connection import get_connection_kwargs

    inserted = 0
    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        for suggestion in suggestions:
            metadata = dict(suggestion.get("metadata") or {})
            suggestion_id = metadata.get("suggestion_id")
            if not suggestion_id:
                refs = ",".join(suggestion.get("candidate_refs") or [])
                suggestion_id = f"{metadata.get('suggestion_kind')}:{suggestion['company_code']}:{refs}"
                metadata["suggestion_id"] = suggestion_id
            for raw_name, candidate_ref in zip(
                suggestion.get("raw_names") or [],
                suggestion.get("candidate_refs") or [],
                strict=False,
            ):
                row_metadata = {
                    **metadata,
                    "candidate_ref": candidate_ref,
                    "raw_name": raw_name,
                }
                conn.execute(
                    'DELETE FROM derived_layer.company_aliases '
                    'WHERE review_status = %s AND source_file = %s '
                    '  AND "申請人代碼" = %s '
                    r"  AND alias_lookup_key = lower(regexp_replace(btrim(%s), '\s+', ' ', 'g'))",
                    (
                        "ai_suggested",
                        "ai:company_normalization_suggestion",
                        suggestion["company_code"],
                        raw_name,
                    ),
                )
                conn.execute(
                    'INSERT INTO derived_layer.company_aliases '
                    '("申請人代碼", "公司中文名稱", "正規化名稱", "別稱", '
                    ' source_file, source_type, review_status, wips_metadata_json) '
                    "VALUES (%s, %s, %s, %s, %s, 'ai_suggested', 'ai_suggested', %s)",
                    (
                        suggestion["company_code"],
                        suggestion.get("zh_name"),
                        suggestion.get("normalized_name"),
                        raw_name,
                        "ai:company_normalization_suggestion",
                        Jsonb(row_metadata),
                    ),
                )
                inserted += 1
        if inserted:
            notify_company_aliases_changed(conn, action="ingest_suggestions")
        conn.commit()
    return {"inserted": inserted}


def list_company_normalization_suggestions(
    *,
    limit: int = 100,
    offset: int = 0,
    connect_kwargs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """列出公司正規化待審建議；前端用 metadata 顯示證據，不顯示 raw JSON。"""
    import psycopg
    from psycopg.rows import dict_row

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs()), row_factory=dict_row) as conn:
        items = conn.execute(
            _LIST_COMPANY_NORMALIZATION_SUGGESTIONS_SQL,
            {"limit": limit, "offset": offset},
        ).fetchall()
        total = conn.execute(_COUNT_COMPANY_NORMALIZATION_SUGGESTIONS_SQL).fetchone()["total"]
    return {"items": [dict(row) for row in items], "total": int(total)}


def clear_company_normalization_suggestions(
    suggestion_ids: list[int],
    *,
    connect_kwargs: dict[str, Any] | None = None,
) -> int:
    """刪掉已確認的待審列；略過的建議不走這裡，仍留待稍後處理。"""
    if not suggestion_ids:
        return 0
    import psycopg

    from backend.app.db.connection import get_connection_kwargs

    with psycopg.connect(**(connect_kwargs or get_connection_kwargs())) as conn:
        cur = conn.execute(
            "DELETE FROM derived_layer.company_aliases "
            "WHERE id = ANY(%s) AND review_status = 'ai_suggested' "
            "  AND source_file = 'ai:company_normalization_suggestion'",
            (suggestion_ids,),
        )
        deleted = int(cur.rowcount or 0)
        if deleted:
            notify_company_aliases_changed(conn, action="review_suggestions")
        conn.commit()
    return deleted


def confirm_company_normalization_suggestions(
    decisions: list[Any],
    *,
    source_label: str = "display_name_curation:company_normalization_review",
) -> dict[str, Any]:
    """把人工確認的 AI 建議轉成 confirmed mapping；正式寫入仍委派唯一 writer。"""
    confirm_ids = [int(getattr(item, "suggestion_id")) for item in decisions if getattr(item, "action") == "confirm"]
    skip_ids = [int(getattr(item, "suggestion_id")) for item in decisions if getattr(item, "action") == "skip"]
    skipped = [item for item in decisions if getattr(item, "action") == "skip"]
    if not confirm_ids:
        cleared = clear_company_normalization_suggestions(skip_ids)
        return {"confirmed": 0, "skipped": len(skipped), "written": {"inserted": 0}, "drafts_cleared": cleared}

    drafts = list_company_normalization_suggestions(limit=1000, offset=0)["items"]
    drafts_by_id = {int(row["id"]): row for row in drafts}
    missing = [sid for sid in confirm_ids if sid not in drafts_by_id]
    if missing:
        raise ValueError(f"company normalization suggestion not found or stale: {missing[0]}")
    targets_by_code = {
        str(row["code"]): row
        for row in list_company_normalization_targets()
    }

    mapping: dict[str, dict[str, Any]] = {}
    for decision in decisions:
        if getattr(decision, "action") != "confirm":
            continue
        draft = drafts_by_id[int(getattr(decision, "suggestion_id"))]
        target_code = clean_text(getattr(decision, "target_code", None))
        target = targets_by_code.get(target_code) if target_code else None
        if target_code and target is None:
            raise ValueError(f"unknown target_code: {target_code}")
        code = target_code or clean_text(draft.get("company_code"))
        if not code:
            raise ValueError("suggestion missing company_code")
        zh = (
            clean_text(getattr(decision, "zh_name", None))
            or clean_text(target.get("zh_name") if target else None)
            or clean_text(draft.get("suggested_zh_name"))
        )
        en = (
            clean_text(getattr(decision, "normalized_name", None))
            or clean_text(target.get("normalized_name") if target else None)
            or clean_text(draft.get("suggested_normalized_name"))
        )
        if not (zh or en):
            raise ValueError("confirmed suggestion requires zh_name or normalized_name")
        entry = mapping.setdefault(code, {"zh_name": zh, "normalized_name": en, "aliases": []})
        # 同公司多筆一起確認時，最後一筆使用者編輯的公司名視為公司層 canonical。
        entry["zh_name"] = zh or entry.get("zh_name")
        entry["normalized_name"] = en or entry.get("normalized_name")
        raw_name = clean_text(draft.get("raw_name"))
        if raw_name:
            entry["aliases"].append(raw_name)

    written = apply_confirmed_display_names(mapping, source_label)
    cleared = clear_company_normalization_suggestions(confirm_ids + skip_ids)
    return {
        "confirmed": len(confirm_ids),
        "skipped": len(skipped),
        "written": written,
        "drafts_cleared": cleared,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import company alias table into derived_layer.company_aliases.")
    parser.add_argument("path", type=Path, help="Path to company alias CSV/XLSX.")
    parser.add_argument("--dry-run", action="store_true", help="Read and normalize without writing to database.")
    args = parser.parse_args()
    summary = import_company_aliases(args.path, dry_run=args.dry_run)
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
