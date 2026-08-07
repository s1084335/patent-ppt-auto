from __future__ import annotations

import argparse
import csv
import hashlib
import json
import posixpath
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

from backend.app.mappings.wips import (
    APPLICATION_DATE_FIELD,
    ATTRIBUTE_FIELD_COLUMNS,
    ATTRIBUTE_FIELDS,
    MAPPING_VERSION,
    PATENT_FIELDS,
    PEOPLE_FIELD_COLUMNS,
    PEOPLE_GROUPS,
    PEOPLE_STAT_FIELDS,
    PUBLICATION_DATE_FIELDS,
    SOURCE_SYSTEM,
    canonical_field_name,
)
from backend.app.derived.company_alias_importer import (
    build_people_pairs,
    people_value,
    register_known_code_variants,
)
from backend.app.transforms.dates import parse_date, year_from_date
from backend.app.transforms.text import clean_long_text, clean_text, value_to_text
from backend.app.transforms.patent_numbers import (
    APPLICATION_NUMBER_TRANSFORMED,
    UNEXAMINED_PUBLICATION_NUMBER_TRANSFORMED,
    transform_patent_number,
    transformed_number_fields,
)

GRANT_PUBLICATION_NUMBER_FIELD = "授权公告号"
UNEXAMINED_PUBLICATION_NUMBER_FIELD = "未审查的公开号"
EXAMINED_PUBLICATION_NUMBER_FIELD = "审查的公告号"
APPLICATION_NUMBER_FIELD = "申请号"
IDENTIFIER_SOURCE_FIELDS = (
    GRANT_PUBLICATION_NUMBER_FIELD,
    EXAMINED_PUBLICATION_NUMBER_FIELD,
    UNEXAMINED_PUBLICATION_NUMBER_FIELD,
    APPLICATION_NUMBER_FIELD,
)
PATENT_IDENTIFIER_LOOKUP_ORDER = (
    "授權公告號",
    "審查的公告號",
    UNEXAMINED_PUBLICATION_NUMBER_TRANSFORMED,
)
# patent_people 要寫入的 WIPS 來源欄（順序即 INSERT 欄序，`dict.fromkeys` 去重保序）。
# ⚠ 必須是 `PEOPLE_GROUPS` ∪ `PEOPLE_STAT_FIELDS` 兩者聯集：發明人數／申請人數是
# **人員統計**而非某個角色的欄位，放進 `PEOPLE_GROUPS["applicant"]` 會讓
# 「一組＝一種角色的欄位集」的語意走樣，故 mapping 端刻意分開宣告（見 wips.py）。
# 這裡只漏掉統計那組的話，症狀是「欄位存在、匯入不報錯、值永遠 NULL」——不會有錯誤訊息。
PEOPLE_FIELDS = tuple(dict.fromkeys(
    [field for fields in PEOPLE_GROUPS.values() for field in fields.values()]
    + list(PEOPLE_STAT_FIELDS)
))
# core_layer.patents 的代表圖欄（0026 起 bytea）；0031 起語意降為「最新階段圖的快取」，
# 完整保存在 core_layer.patent_figures。值來自 xlsx 內嵌浮動圖片，非儲存格文字。
FIGURE_COLUMN = "主附圖"
# core_layer.patents.document_kind 的欄名（PATENT_FIELDS 的目標名）；代表圖依此分階段保存。
DOCUMENT_KIND_COLUMN = "document_kind"
# document_kind 缺值時的佔位（patent_figures.document_kind 為 NOT NULL，不得因缺欄丟圖）。
UNKNOWN_DOCUMENT_KIND = "UNKNOWN"
# 文獻階段優先序：rank 越大越後期，主表快取取最大者（B 審定公告 > A 早期公開）。
# 未知類型 rank=0，入庫保存但不覆蓋任何已知階段的快取；新增制度只加對照，不改邏輯。
# 不得改用 ORDER BY document_kind DESC——字母序 ≠ 階段序（WIPS 另有 A1/A2/B1/B2/U/Y）。
_KIND_RANK: dict[str, int] = {"A": 1, "B": 2}


def figure_kind_rank(document_kind: str | None) -> int:
    """取文獻種類的階段 rank；未知或缺值回 0（保守，不覆蓋已知階段）。

    先查完整值（涵蓋 "A"／"B"），再退回首字母（涵蓋各國制度的 A1/A2/B1/B2 等同階段變體）。
    """
    if not document_kind:
        return 0
    kind = str(document_kind).strip().upper()
    if kind in _KIND_RANK:
        return _KIND_RANK[kind]
    return _KIND_RANK.get(kind[:1], 0)
CONFLICT_RESOLUTION_STRATEGY = "incoming_source_priority"

# CLI importer 支援的全部來源副檔名（含 .mdb）；與 load_source_rows 的分派一致。
# Web 上傳/worker 另有較窄白名單（不含 .mdb）於 importers.import_paths.WEB_IMPORT_SUFFIXES。
SUPPORTED_IMPORT_SUFFIXES = (".xlsx", ".csv", ".txt", ".xml", ".mdb")

# update_patent_changed_fields 的 (欄位, 參數名) 對照：用來逐欄組「差異即更新」CASE 與
# WHERE guard，區分 no-op 與實際更新。不寫死欄名清單——SET/guard 皆由本表生成，
# 日後增減欄位只改此表即可。
_UPDATE_COLUMN_PARAMS: tuple[tuple[str, str], ...] = (
    ('"授權公告號"', "授權公告號"),
    ('"審查的公告號"', "審查的公告號"),
    ('"未審查的公開號"', "未審查的公開號"),
    ('"申請號"', "申請號"),
    ("country_code", "country_code"),
    ("database_name", "database_name"),
    ("document_kind", "document_kind"),
    ("patent_type", "patent_type"),
    ("publication_date", "publication_date"),
    ("publication_year", "publication_year"),
    ("application_date", "application_date"),
    ("application_year", "application_year"),
    ("title", "title"),
    ("title_original", "title_original"),
    ("abstract", "abstract"),
    ('"權利要求的項數"', "claim_count"),
    ('"所有權利要求[JP,KR,CN]"', "all_claims"),
    ('"主權項"', "main_claim"),
    ('"主權項(原文)"', "main_claim_original"),
    ('"獨立項數量[KR,JP,US,CN,EP,IN]"', "independent_claim_count"),
    ('"獨立項[KR,JP,US,CN,EP,IN]"', "independent_claims"),
    ('"獨立項(原文)[KR,JP,CN,EP]"', "independent_claims_original"),
    ('"效果 摘要[US,EP,PCT,JP,KR,CN,TW]"', "effect_summary"),
    ('"Orig. CPC(Main)"', "orig_cpc_main"),
    ('"Orig. IPC(Main)"', "orig_ipc_main"),
    ('"Curr. CPC(Main)"', "curr_cpc_main"),
    ('"Curr. IPC(Main)"', "curr_ipc_main"),
    ("legal_status", "legal_status"),
    ('"WIPS同族ID"', "WIPS同族ID"),
    # ── 欄位重分類（0046，2026-08-06）：原本落在 patent_attributes 的 14 欄 ──
    # 🔴 參數名**不得含括號**：psycopg 解析 `%(name)s` 時以第一個 `)` 為結尾，
    # `%(摘要(原文)_cmp)s` 會被截斷，整句 UPDATE 拋 ProgrammingError。
    # ⚠ 這正是上面那批舊欄一律用英文別名的原因——不是歷史包袱，是這條限制的產物。
    # （0046 初版誤判為包袱、直接拿欄名當參數名，實機跑真 DB 才炸出來；
    #   字串契約測試看不到這種錯，已補 `test_param_names_are_psycopg_safe`。）
    ('"摘要(原文)"', "abstract_original"),
    # ⚠ 不可叫 publication_date——那個參數名已被「解析後的公開日期」佔用（見上）。
    ('"未審查的公開日"', "unexamined_publication_date"),
    ('"授權公告日"', "grant_date"),
    ('"優先權號"', "priority_number"),
    ('"優先權國家"', "priority_country"),
    ('"優先權日"', "priority_date"),
    ('"詳細查看連結(登入)"', "detail_url"),
    ('"文圖像文件(PDF)連結"', "pdf_url"),
    ('"WIPS同族各國家文獻數量(申請為準)"', "family_country_doc_counts"),
    ('"EPC有效國家[EP]"', "epc_valid_countries"),
    ('"EPC無效國家[EP]"', "epc_invalid_countries"),
    ('"(F1)引用文獻數"', "f1_citation_count"),
    ('"(B1)引用文獻數"', "b1_citation_count"),
    ('"解決課題 摘要[US,EP,PCT,JP,KR,CN,TW]"', "problem_summary"),
)


def build_patent_params(patent: dict[str, Any]) -> dict[str, Any]:
    """把 normalize_record 的 patent dict 補上 `_UPDATE_COLUMN_PARAMS` 需要的別名。

    `patent` 的 key 是**目標欄名**（`normalize_record` 以 `PATENT_FIELDS` 的 target
    為 key），而 SQL 用的是參數名——兩者在有別名的欄位上不同，必須補齊。

    ⚠ 0046 前這裡是一段**手寫的 12 行別名 dict**，等於 `_UPDATE_COLUMN_PARAMS`
    的第三份落點。漏補一個的症狀是靜默的：`.get(param)` 回 None → 該欄永遠寫入
    NULL，不報錯。改為由 `_UPDATE_COLUMN_PARAMS` 推導後，加欄只改那一張表。
    `setdefault` 讓已算好的衍生值（publication_date／application_year 等）不被蓋掉。
    """
    params = {**patent}
    for column, param in _UPDATE_COLUMN_PARAMS:
        params.setdefault(param, patent.get(column.strip('"')))
    return params


def _patent_insert_sql() -> str:
    """由 `_UPDATE_COLUMN_PARAMS` 生成 INSERT——欄位清單不再抄第二份。

    ⚠ 0046 前 INSERT 是一段寫死的 SQL 字面，與 `_UPDATE_COLUMN_PARAMS` 內容
    **一字不差地重複**（實測 29 欄、同序、同參數名）。兩份各自演進的後果是靜默的：
    新欄只加進其中一份時，另一條路徑照跑不報錯，值卻永遠是 NULL。
    收成一處後，加欄只改 `_UPDATE_COLUMN_PARAMS`，insert／update 同時跟上。
    """
    columns = ", ".join(column for column, _ in _UPDATE_COLUMN_PARAMS)
    values = ", ".join(f"%({param})s" for _, param in _UPDATE_COLUMN_PARAMS)
    return f"INSERT INTO patents (\n            {columns}\n        )\n        VALUES (\n            {values}\n        )\n        RETURNING id"


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_source_rows(path: Path) -> tuple[list[str], str, list[dict[str, Any]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_xlsx_rows(path)
    if suffix == ".csv":
        return load_delimited_rows(path, source_name=path.name)
    if suffix == ".txt":
        return load_delimited_rows(path, source_name=path.name)
    if suffix == ".xml":
        return load_xml_rows(path)
    if suffix == ".mdb":
        return load_mdb_rows(path)
    raise ValueError(f"Unsupported WIPS export format: {path.suffix}")


def load_xlsx_rows(path: Path) -> tuple[list[str], str, list[dict[str, Any]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    selected_sheet = select_patent_sheet(workbook)
    worksheet = workbook[selected_sheet]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records = []
    for row_number, row in enumerate(rows, start=2):
        raw = {}
        has_value = False
        for header, value in zip(headers, row):
            if not header:
                continue
            raw[header] = value
            has_value = has_value or value not in (None, "")
        if has_value:
            raw["_row_number"] = row_number
            records.append(raw)
    # 內嵌代表圖：以 zipfile 直接解析 drawing XML 取列號與圖片（不再全載 workbook；
    # B-1 維運修復），大檔不 hang。回 {列號: (圖片 bytes, 該列圖片張數)}；
    # 張數 >1 由 import_wips_file 轉成 summary 警告。
    embedded = extract_embedded_images(path, selected_sheet)
    for raw in records:
        found = embedded.get(raw["_row_number"])
        if found is None:
            continue
        blob, count = found
        raw[EMBEDDED_IMAGE_KEY] = blob
        if count > 1:
            raw[FIGURE_WARNINGS_KEY] = _MULTI_IMAGE_WARNING.format(
                row_number=raw["_row_number"], count=count
            )
    return sheet_names, selected_sheet, records, headers


# 內嵌圖在 record 內的暫存鍵；與 _row_number 同屬非來源欄，不寫進 raw_records.raw_data。
EMBEDDED_IMAGE_KEY = "_embedded_image"
FIGURE_WARNINGS_KEY = "_figure_warning"
# 同一資料列偵測到多張圖時的警告訊息（取第一張，其餘明確記錄不靜默丟棄）。
_MULTI_IMAGE_WARNING = "第 {row_number} 列偵測到 {count} 張內嵌圖，僅取第一張（其餘未入庫）"
# 同一 (patent_id, document_kind) 在同批出現多張圖：PK 只留一列，須明確記錄不靜默丟棄。
_DUPLICATE_KIND_WARNING = (
    "專利 {label} 的文獻種類「{kind}」在本批出現 {count} 張圖，僅保留第一張（其餘未入庫）"
)
# 未知文獻種類：仍入庫保存，但 rank=0 不覆蓋已知階段的主表快取。
_UNKNOWN_KIND_WARNING = "專利 {label} 的文獻種類「{kind}」未列於階段對照，圖已入庫但不作為最新版快取"
# 文獻種類缺值：以 UNKNOWN 落庫，不因缺欄丟圖。
_MISSING_KIND_WARNING = "專利 {label} 缺文獻種類，圖以「{placeholder}」入庫（不作為最新版快取）"


# drawing XML／關聯檔命名空間無關解析用的原子標籤名（去掉 xdr:/r: 等前綴後比對）。
# 真檔（Excel/WIPS 匯出）多用 xdr: 前綴＋twoCellAnchor；openpyxl 產的檔用預設命名空間
# ＋oneCellAnchor。兩者的 <from><row>、<blip r:embed>、<Relationship> 結構一致，
# 故一律以 local name（去命名空間）比對，不寫死任一種前綴。
_DRAWING_ROW_TAG = "row"
_DRAWING_FROM_TAG = "from"
_DRAWING_BLIP_TAG = "blip"
_DRAWING_EMBED_ATTR = "embed"
_RELATIONSHIP_TAG = "Relationship"


def _local_name(tag: str) -> str:
    """去掉 XML 標籤的命名空間前綴（`{ns}row`→`row`），供命名空間無關比對。"""
    return tag.rsplit("}", 1)[-1]


def _local_attr(attrib: dict[str, str], name: str) -> str | None:
    """以 local name 取屬性值（`r:embed` 序列化後為 `{ns}embed`），找不到回 None。"""
    for key, value in attrib.items():
        if _local_name(key) == name:
            return value
    return None


def _parse_relationships(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    """讀一份 .rels，回 {Id: 解析後的 zip 內部路徑}；缺檔或解析失敗回空 dict。

    Target 可能是絕對（/xl/...）或相對（../media/image1.png）；以 rels 檔所在目錄的
    上一層（package part 慣例）為基準解析成 zip 內的正規化路徑，供 zf.read 直接取用。
    """
    try:
        data = zf.read(rels_path)
    except KeyError:
        return {}
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return {}
    # rels 檔位於 <part 所在目錄>/_rels/<part>.rels；相對 Target 以 part 所在目錄為基準。
    base_dir = posixpath.dirname(posixpath.dirname(rels_path))
    mapping: dict[str, str] = {}
    for element in root:
        if _local_name(element.tag) != _RELATIONSHIP_TAG:
            continue
        rel_id = element.attrib.get("Id")
        target = element.attrib.get("Target")
        if not rel_id or not target:
            continue
        if target.startswith("/"):
            resolved = target.lstrip("/")
        else:
            resolved = posixpath.normpath(posixpath.join(base_dir, target))
        mapping[rel_id] = resolved
    return mapping


def _resolve_sheet_part(zf: zipfile.ZipFile, sheet_name: str) -> str | None:
    """由工作表顯示名解析出其 zip 內的 worksheet part 路徑（xl/worksheets/sheetN.xml）。

    鏈路：workbook.xml 的 <sheet name r:id> → workbook.xml.rels 的 rId→Target。
    sheetN 的 N 與工作表順序無必然關係，故必經 rels 解析，不可用 index 推。
    """
    try:
        workbook_xml = zf.read("xl/workbook.xml")
    except KeyError:
        return None
    try:
        root = ElementTree.fromstring(workbook_xml)
    except ElementTree.ParseError:
        return None
    rels = _parse_relationships(zf, "xl/_rels/workbook.xml.rels")
    for element in root.iter():
        if _local_name(element.tag) != "sheet":
            continue
        if element.attrib.get("name") != sheet_name:
            continue
        rel_id = _local_attr(element.attrib, "id")
        if rel_id and rel_id in rels:
            return rels[rel_id]
    return None


def _resolve_drawing_part(zf: zipfile.ZipFile, sheet_part: str) -> str | None:
    """由 worksheet part 找出其 drawing part 路徑（xl/drawings/drawingN.xml）。

    鏈路：sheetN.xml 的 <drawing r:id> → sheetN.xml.rels 的 rId→Target。無 drawing 回 None。
    """
    try:
        sheet_xml = zf.read(sheet_part)
    except KeyError:
        return None
    try:
        root = ElementTree.fromstring(sheet_xml)
    except ElementTree.ParseError:
        return None
    drawing_rel_id: str | None = None
    for element in root.iter():
        if _local_name(element.tag) == "drawing":
            drawing_rel_id = _local_attr(element.attrib, "id")
            if drawing_rel_id:
                break
    if not drawing_rel_id:
        return None
    sheet_dir = posixpath.dirname(sheet_part)
    rels_path = f"{sheet_dir}/_rels/{posixpath.basename(sheet_part)}.rels"
    rels = _parse_relationships(zf, rels_path)
    return rels.get(drawing_rel_id)


def _anchor_row_number(anchor: ElementTree.Element) -> int | None:
    """由一個 anchor 元素取 1-based 列號：<from><row>{0-based}</row></from> + 1。

    oneCellAnchor／twoCellAnchor 皆有 <from><row>；命名空間無關比對。缺 from/row 回 None。
    """
    for child in anchor:
        if _local_name(child.tag) != _DRAWING_FROM_TAG:
            continue
        for marker in child:
            if _local_name(marker.tag) == _DRAWING_ROW_TAG and marker.text is not None:
                try:
                    return int(marker.text.strip()) + 1
                except ValueError:
                    return None
    return None


def _anchor_embed_id(anchor: ElementTree.Element) -> str | None:
    """由一個 anchor 元素取其圖片關聯 Id（<blip r:embed="rId"/>）；無圖回 None。"""
    for element in anchor.iter():
        if _local_name(element.tag) == _DRAWING_BLIP_TAG:
            embed = _local_attr(element.attrib, _DRAWING_EMBED_ATTR)
            if embed:
                return embed
    return None


def extract_embedded_images(path: Path, sheet_name: str) -> dict[int, tuple[bytes, int]]:
    """取出 xlsx 指定工作表的內嵌圖片，回 {1-based 列號: (第一張圖 bytes, 該列圖片張數)}。

    WIPS 匯出的「主附圖」是錨在資料列的**浮動圖片物件**，與儲存格文字值無關
    （儲存格值為單一空白字元）。**改用 zipfile 直接解析 drawing XML**，不再
    load_workbook 全載整個 workbook（B-1 維運修復：read_only=False 全載會解析全部
    drawing/樣式/公式，53.8MB 檔 hang/爆記憶體）。

    對映與舊版完全一致：drawing XML 的 <from><row>（0-based）+1 = 1-based 列號，
    對上 load_xlsx_rows 的 `_row_number`；圖片 bytes 直接由 xl/media/imageN.xxx 讀出。

    解析鏈：workbook.xml→sheet part→drawingN.xml（列號＋blip embed id）＋
    drawingN.xml.rels（embed id→media 路徑）。命名空間無關比對，兼容 Excel 產的
    xdr:/twoCellAnchor 與 openpyxl 產的預設命名空間/oneCellAnchor。

    通用性：不假設圖片在哪一欄、不假設每列都有圖、不假設圖片數等於資料列數。
    完全沒有內嵌圖、檔案結構不支援 drawing、非 zip 或損毀時回空 dict，匯入照常進行。
    同一列多張圖時回傳第一張並附上張數，由呼叫端轉成警告，不靜默丟棄。
    """
    images: dict[int, bytes] = {}
    counts: dict[int, int] = {}
    try:
        with zipfile.ZipFile(path) as zf:
            sheet_part = _resolve_sheet_part(zf, sheet_name)
            if not sheet_part:
                return {}
            drawing_part = _resolve_drawing_part(zf, sheet_part)
            if not drawing_part:
                return {}
            try:
                drawing_xml = zf.read(drawing_part)
            except KeyError:
                return {}
            drawing_root = ElementTree.fromstring(drawing_xml)
            drawing_dir = posixpath.dirname(drawing_part)
            drawing_rels = _parse_relationships(
                zf,
                f"{drawing_dir}/_rels/{posixpath.basename(drawing_part)}.rels",
            )
            # 逐 anchor（文件序＝Excel 錨點序）取列號與圖片；同列多張時 setdefault
            # 保留先到者（第一張），counts 累加該列張數——與舊版 openpyxl 走訪語意一致。
            for anchor in drawing_root:
                row_number = _anchor_row_number(anchor)
                if row_number is None:
                    continue
                embed_id = _anchor_embed_id(anchor)
                if not embed_id:
                    continue
                media_path = drawing_rels.get(embed_id)
                if not media_path:
                    continue
                try:
                    blob = zf.read(media_path)
                except KeyError:
                    continue
                if not blob:
                    continue
                counts[row_number] = counts.get(row_number, 0) + 1
                images.setdefault(row_number, blob)
    except (zipfile.BadZipFile, ElementTree.ParseError, OSError):
        # 非 zip、加密、損毀 drawing／關聯：圖片屬加值資料，取不到就當作無圖，
        # 不得讓整批匯入失敗（沿舊版容錯）。
        return {}
    return {row: (blob, counts[row]) for row, blob in images.items()}


def load_delimited_rows(path: Path, source_name: str) -> tuple[list[str], str, list[dict[str, Any]], list[str]]:
    """讀 csv/txt 逐列串流解析（2026-07-26 修 bug1/bug2）。

    bug1（引號內換行）：舊版 `csv.reader(text.splitlines(), ...)` 先 splitlines 會吃掉
    引號內儲存格的換行、破壞列結構。改用 csv.reader 直接吃**檔案物件**（newline="" 為
    官方要求，讓 csv 模組自己正確處理引號內換行）；順帶解掉整檔全載（逐列串流）。
    bug2（Sniffer 只看前 8KB）：擴大取樣（前 64KB 且多列），sniff 失敗/低信心時
    退到副檔名預設（.txt→tab、.csv→逗號）。
    """
    encoding = _detect_text_encoding(path)
    # 先讀較大樣本判分隔符（bug2：8KB→64KB，前段無代表性時較不易猜錯）。
    with open(path, "r", encoding=encoding, newline="") as f:
        sample = f.read(65536)
    try:
        dialect: Any = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel_tab if path.suffix.lower() == ".txt" else csv.excel

    # bug1：csv.reader 吃檔案物件（不 splitlines），引號內換行由 csv 模組原生正確處理。
    records = []
    headers: list[str] = []
    with open(path, "r", encoding=encoding, newline="") as f:
        reader = csv.reader(f, dialect)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(reader)]
        except StopIteration:
            return [source_name], source_name, [], []
        for row_number, row in enumerate(reader, start=2):
            raw = row_to_record(headers, row)
            if record_has_value(raw):
                raw["_row_number"] = row_number
                records.append(raw)
    return [source_name], source_name, records, headers


def _detect_text_encoding(path: Path) -> str:
    """偵測可用編碼（沿 read_text_with_fallback 的順序），供串流開檔用。

    讀前 64KB 逐一試解碼，回第一個成功的編碼；全失敗回 utf-8（開檔時以 errors 容錯）。
    這樣改串流讀檔時不犧牲既有的多編碼容錯。
    """
    head = path.read_bytes()[:65536]
    for encoding in ("utf-8-sig", "utf-16", "big5", "cp950", "gb18030"):
        try:
            head.decode(encoding)
            return encoding
        except UnicodeError:
            continue
    return "utf-8"


def read_text_with_fallback(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-16", "big5", "cp950", "gb18030"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def load_xml_rows(path: Path) -> tuple[list[str], str, list[dict[str, Any]], list[str]]:
    tree = ElementTree.parse(path)
    root = tree.getroot()
    candidates: list[tuple[str, int, Any, dict[str, Any]]] = []
    for element in root.iter():
        if element is root:
            continue
        flattened = flatten_xml_element(element)
        if not flattened:
            continue
        score = patent_marker_score(flattened)
        if score:
            candidates.append((local_name(element.tag), score, element, flattened))

    if candidates:
        best_tag, _, _, _ = max(candidates, key=lambda item: (item[1], sum(1 for c in candidates if c[0] == item[0])))
        selected = [(element, flattened) for tag, _, element, flattened in candidates if tag == best_tag]
        selected_name = best_tag
    else:
        selected = [(element, flatten_xml_element(element)) for element in list(root) if flatten_xml_element(element)]
        selected_name = local_name(root.tag)

    headers = ordered_headers([flattened for _, flattened in selected])
    records = []
    for row_number, (_, flattened) in enumerate(selected, start=1):
        raw = {header: flattened.get(header) for header in headers}
        if record_has_value(raw):
            raw["_row_number"] = row_number
            records.append(raw)
    return [selected_name], selected_name, records, headers


def load_mdb_rows(path: Path) -> tuple[list[str], str, list[dict[str, Any]], list[str]]:
    try:
        import pyodbc
    except ImportError as exc:
        raise RuntimeError("MDB import requires pyodbc and a Microsoft Access ODBC driver.") from exc

    connection_string = (
        r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
        f"DBQ={path};"
    )
    with pyodbc.connect(connection_string) as conn:
        cursor = conn.cursor()
        table_names = [
            row.table_name
            for row in cursor.tables(tableType="TABLE")
            if not str(row.table_name).startswith("MSys")
        ]
        if not table_names:
            return [path.name], path.name, [], []
        selected_table = select_patent_table(conn, table_names)
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM [{selected_table}]")
        headers = [column[0] for column in cursor.description]
        records = []
        for row_number, row in enumerate(cursor.fetchall(), start=1):
            raw = row_to_record(headers, list(row))
            if record_has_value(raw):
                raw["_row_number"] = row_number
                records.append(raw)
    return table_names, selected_table, records, headers


def select_patent_table(conn: Any, table_names: list[str]) -> str:
    best_table = table_names[0]
    best_score = -1
    for table_name in table_names:
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM [{table_name}] WHERE 1=0")
        headers = {canonical_field_name(column[0]) for column in cursor.description}
        score = len({"申请号", "标题", "申请日"} & headers)
        if score > best_score:
            best_score = score
            best_table = table_name
    return best_table


def row_to_record(headers: list[str], row: list[Any]) -> dict[str, Any]:
    raw = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        raw[header] = row[index] if index < len(row) else None
    return raw


# raw record 中的非來源欄（內部用），不參與判空、正規化與 raw_records.raw_data 寫入。
_INTERNAL_RECORD_KEYS = ("_row_number", EMBEDDED_IMAGE_KEY, FIGURE_WARNINGS_KEY)


def record_has_value(raw: dict[str, Any]) -> bool:
    return any(
        value_to_text(value) is not None
        for key, value in raw.items()
        if key not in _INTERNAL_RECORD_KEYS
    )


def ordered_headers(records: list[dict[str, Any]]) -> list[str]:
    headers = []
    seen = set()
    for record in records:
        for key in record:
            if key not in seen:
                seen.add(key)
                headers.append(key)
    return headers


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def flatten_xml_element(element: Any) -> dict[str, Any]:
    flattened = {local_name(key): value for key, value in element.attrib.items()}
    children = list(element)
    if not children:
        text = (element.text or "").strip()
        return {local_name(element.tag): text} if text else flattened
    for child in children:
        child_name = xml_field_name(child)
        child_children = list(child)
        if child_children:
            child_values = flatten_xml_element(child)
            for key, value in child_values.items():
                flattened.setdefault(key, value)
        else:
            flattened.setdefault(child_name, (child.text or "").strip() or None)
            for attr_key, attr_value in child.attrib.items():
                if attr_key in XML_FIELD_NAME_ATTRIBUTES:
                    continue
                flattened.setdefault(f"{child_name}_{local_name(attr_key)}", attr_value)
    return flattened


XML_FIELD_NAME_ATTRIBUTES = ("name", "field", "label", "title")


def xml_field_name(element: Any) -> str:
    for attribute_name in XML_FIELD_NAME_ATTRIBUTES:
        value = element.attrib.get(attribute_name)
        if value:
            return str(value).strip()
    return local_name(element.tag)


def patent_marker_score(record: dict[str, Any]) -> int:
    canonical_headers = {canonical_field_name(key) for key in record}
    return len({"申请号", "标题", "申请日"} & canonical_headers)


def select_patent_sheet(workbook) -> str:
    required_markers = {"申请号", "标题", "申请日"}
    best_sheet = workbook.sheetnames[0]
    best_score = -1
    for worksheet in workbook.worksheets:
        try:
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            continue
        headers = {canonical_field_name(str(value)) for value in first_row if value is not None}
        score = len(required_markers & headers)
        if score > best_score:
            best_score = score
            best_sheet = worksheet.title
    return best_sheet


def normalize_record(raw: dict[str, Any]) -> dict[str, Any]:
    canonical_raw = canonicalize_record(raw)
    application_date = parse_date(canonical_raw.get(APPLICATION_DATE_FIELD))
    publication_date = first_parsed_date(canonical_raw, PUBLICATION_DATE_FIELDS)

    # 文字欄一律走 clean_long_text；主附圖為二進位（bytea），不得經文字清洗，故排除後另填。
    patent = {
        target: clean_long_text(canonical_raw.get(source))
        for source, target in PATENT_FIELDS.items()
        if target != FIGURE_COLUMN
    }
    patent.update(transformed_number_fields(patent))
    # 🔴 兩個狀態欄的優先序**唯一定義處**（2026-08-07 護欄一）：既有來源
    # `状态[US,JP,KR,CN,EP,CA,AU]` 非空優先，空才取「專利狀態」（全國家欄，
    # canonicalize 後鍵為簡體「专利状态」）——既有國家行為零改變，
    # 只有 TW（前者恆空）由後者補值。
    if not (patent.get("legal_status") or "").strip():
        alt_status = clean_long_text(canonical_raw.get("专利状态"))
        if alt_status and alt_status.strip():
            patent["legal_status"] = alt_status
    patent["publication_date"] = publication_date
    patent["publication_year"] = year_from_date(publication_date)
    patent["application_date"] = application_date
    patent["application_year"] = year_from_date(application_date)

    return {
        "patent": patent,
        "people": normalize_people(canonical_raw),
        "attributes": normalize_attributes(canonical_raw),
    }


def first_parsed_date(raw: dict[str, Any], fields: list[str]) -> Any:
    for field in fields:
        parsed = parse_date(raw.get(field))
        if parsed:
            return parsed
    return None


def canonicalize_record(raw: dict[str, Any]) -> dict[str, Any]:
    canonical = {}
    for key, value in raw.items():
        if key in _INTERNAL_RECORD_KEYS:
            canonical[key] = value
            continue
        canonical.setdefault(canonical_field_name(key), value)
    return canonical


def normalize_people(raw: dict[str, Any]) -> dict[str, Any]:
    people = {}
    for source_field in PEOPLE_FIELDS:
        people[source_field] = clean_text(raw.get(source_field))
    return people


def normalize_attributes(raw: dict[str, Any]) -> dict[str, Any]:
    return {source_field: value_to_text(raw.get(source_field)) for source_field in ATTRIBUTE_FIELDS}


def infer_value_type(source_field: str, value: Any) -> str:
    if parse_date(value):
        return "date"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if "链接" in source_field or "link" in source_field.lower():
        return "url"
    return "text"


def import_wips_file(path: Path, dry_run: bool = False) -> dict[str, Any]:
    source_names, selected_source, records, headers = load_source_rows(path)
    normalized = [normalize_record(record) for record in records]
    file_hash = file_sha256(path)
    summary = {
        "source_system": SOURCE_SYSTEM,
        "mapping_version": MAPPING_VERSION,
        "file": str(path),
        "file_format": path.suffix.lstrip(".").lower(),
        "file_hash": file_hash,
        "source_names": source_names,
        "selected_source": selected_source,
        "headers": len(headers),
        "records": len(records),
        "normalized_records": len(normalized),
    }
    if dry_run:
        return summary

    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("psycopg is required for database import. Install psycopg[binary].") from exc

    from backend.app.db.connection import get_connection_kwargs

    # 匯入計數；整段在單一 connection／transaction 內，任一步拋錯由 with 區塊 rollback（不 commit）。
    # inserted=新建 patent；matched_existing=識別號命中既有；updated⊆matched_existing（任一欄差異更新）。
    stats = {"inserted": 0, "matched_existing": 0, "updated": 0}
    # 護欄二（2026-08-07）：被人工登錄擋下的狀態覆蓋，逐筆現形交使用者裁決。
    legal_status_conflicts: list[dict[str, Any]] = []
    # 匯入列的 (申請人代表碼, 名稱) 供已知 code 變體即時補入；於 commit 後統一註冊（該函式自管連線）
    variant_pairs: list[tuple[str | None, str | None]] = []
    # 代碼 → WIPS 標準化申請人，供未建組時填英文正式名（規格批次 b）。
    standardized_by_code: dict[str, str] = {}
    # 有 people 資料卻抽不出任何 (代碼, 名稱) 配對的列數——欄名不認得的訊號。
    people_unmatched = 0
    # 本次涉及的 patent_ids（新建＋命中既有），保序去重，供匯入圈 workspace（2026-07-22 定案）。
    touched_patent_ids: list[int] = []
    seen_patent_ids: set[int] = set()
    # 代表圖三元組 (patent_id, document_kind, bytes) 先蒐集、迴圈結束後一次批次寫入，
    # 避免逐列 UPDATE 造成 N+1。(patent_id, kind) 為 patent_figures 主鍵，同鍵重複只留第一張。
    figure_triplets: list[tuple[int, str, bytes]] = []
    # 已收之 (patent_id, kind) → 該鍵在本批出現的張數；>1 者轉成 warning，不靜默丟棄。
    figure_kind_counts: dict[tuple[int, str], int] = {}
    # 各文獻階段入庫張數（如 {"A": 1330, "B": 576}），供 summary 可見性。
    figure_stats: dict[str, int] = {}
    figure_warnings: list[str] = []
    # 主表「主附圖」快取候選：patent_id → (rank, kind, bytes)，全批取 rank 最大者（B > A > 未知）。
    latest_stage_map: dict[int, tuple[int, str, bytes]] = {}
    with psycopg.connect(**get_connection_kwargs()) as conn:
        with conn.cursor() as cur:
            if find_existing_raw_import(cur, file_hash):
                # 同 source_system+source_file_hash 已匯入過（metadata 在 raw_records）：整檔跳過。
                summary["status"] = "skipped_duplicate_file"
                summary["inserted"] = 0
                summary["matched_existing"] = 0
                summary["updated"] = 0
                summary["skipped"] = summary["records"]
                summary["patent_ids"] = []
                return summary
            for raw, item in zip(records, normalized):
                raw_record_id = insert_raw_record(cur, selected_source, raw, file_hash)
                # 去重只靠專利號查找（upsert_patent→find_existing_patent_id）；不再產生 dedupe_key，
                # 無專利號列由 find_existing 回 None 而各自新建，靠 raw_record_id 保持獨立與追溯。
                patent_id = upsert_patent(
                    cur, item["patent"], raw_record_id, stats,
                    legal_status_conflicts=legal_status_conflicts)
                if patent_id not in seen_patent_ids:
                    seen_patent_ids.add(patent_id)
                    touched_patent_ids.append(int(patent_id))
                insert_patent_source(cur, patent_id, raw_record_id)
                replace_people(cur, patent_id, raw_record_id, item["people"])
                replace_attributes(cur, patent_id, raw_record_id, item["attributes"])
                people = item["people"]
                # 五欄配對（2026-07-30 規格 2-6）：原本只取申請人兩欄，
                # 專利權人／受讓人欄的名稱看得見卻不會自動歸戶。改走共用函式，
                # 與待補清單同一套欄位口徑，且 `A | B` 多值會拆開。
                row_pairs = build_people_pairs(people)
                variant_pairs.extend(row_pairs)
                # 🔴 通用性防護（2026-08-03）：**有名稱資料卻一對都抽不到＝欄名不認得**。
                # 本次實機就是這樣：檔案是簡體欄名、`PEOPLE_NAME_CODE_COLUMNS` 寫死繁體，
                # 五欄全部落空 → 治理管線整條空轉，而 summary 只顯示
                # `alias_variants: 0`，看起來像「沒有新變體」而不是「一個都沒掃到」。
                # ⚠ 不預測所有可能的欄名寫法（那是猜），改成**遇到不認得的就出聲**：
                # 日後 WIPS 再換欄名格式，這裡會叫出來而不是靜默失效。
                if people and not row_pairs:
                    people_unmatched += 1
                # 建組時要用的英文正式名：WIPS 標準化申請人（現成資料）。
                # ⚠ 走 `_people_value` 簡繁雙認——原本只寫簡體字面，繁體檔會抓不到，
                # 建組時就填不進英文正式名（同一個 bug 的另一半）。
                _std = clean_text(people_value(people,"標準化申請人"))
                _code = clean_text(people_value(people,"申請人代表碼"))
                if _code and _std:
                    standardized_by_code.setdefault(_code, _std)
                # 同列多圖（load_xlsx_rows 偵測）警告；規格明文要求不得靜默丟棄。
                row_warning = raw.get(FIGURE_WARNINGS_KEY)
                if row_warning:
                    figure_warnings.append(row_warning)
                # 代表圖：只在該列真的有內嵌圖時入列（無圖／無內嵌圖的檔案自然為空清單）。
                blob = raw.get(EMBEDDED_IMAGE_KEY)
                if blob:
                    # 文獻種類已由 normalize_record 依 PATENT_FIELDS 映射進 item["patent"]，
                    # 與寫入 patents.document_kind 的是同一值；缺值以 UNKNOWN 佔位不丟圖。
                    raw_kind = item["patent"].get(DOCUMENT_KIND_COLUMN)
                    kind = str(raw_kind).strip() if raw_kind else ""
                    label = _patent_label(item["patent"], patent_id)
                    if not kind:
                        kind = UNKNOWN_DOCUMENT_KIND
                        figure_warnings.append(
                            _MISSING_KIND_WARNING.format(
                                label=label, placeholder=UNKNOWN_DOCUMENT_KIND
                            )
                        )
                    elif figure_kind_rank(kind) == 0:
                        figure_warnings.append(
                            _UNKNOWN_KIND_WARNING.format(label=label, kind=kind)
                        )
                    key = (int(patent_id), kind)
                    seen_count = figure_kind_counts.get(key, 0) + 1
                    figure_kind_counts[key] = seen_count
                    if seen_count > 1:
                        # 同鍵重複：PK 只容一列，明確記錄後略過（保留先到者，與 upsert 語意一致）。
                        figure_warnings.append(
                            _DUPLICATE_KIND_WARNING.format(
                                label=label, kind=kind, count=seen_count
                            )
                        )
                        continue
                    figure_triplets.append((int(patent_id), kind, blob))
                    figure_stats[kind] = figure_stats.get(kind, 0) + 1
                    # 主表快取候選：同專利取 rank 最大者；rank 相同時保留先到者（確定性）。
                    rank = figure_kind_rank(kind)
                    current = latest_stage_map.get(int(patent_id))
                    if current is None or rank > current[0]:
                        latest_stage_map[int(patent_id)] = (rank, kind, blob)
            # 全部 patent_id 確定後單次批次寫圖（executemany），不逐列往返。
            update_patent_figures(cur, figure_triplets)
            # 再單次批次回寫主表快取（每專利 rank 最大者），同樣不逐列往返。
            update_patent_figure_cache(cur, latest_stage_map)
        conn.commit()
    # 匯入成功：每列一次 upsert，inserted+matched_existing=records、skipped=0；updated 為其中確有差異更新者。
    summary["inserted"] = stats["inserted"]
    summary["matched_existing"] = stats["matched_existing"]
    summary["updated"] = stats["updated"]
    summary["skipped"] = 0
    summary["status"] = "imported"
    summary["legal_status_conflicts"] = legal_status_conflicts
    # 本次涉及專利（新建＋命中既有，去重）供 handler 圈進 workspace；空檔亦回空陣列。
    summary["patent_ids"] = touched_patent_ids
    # 代表圖入庫統計與警告（同列多圖、未知/缺值 kind、同鍵重複）；
    # 無內嵌圖的來源為 0 與空清單，不影響匯入結果。
    summary["figures"] = len(figure_triplets)
    # 各階段入庫張數（如 {"A": 1330, "B": 576}），讓「存了哪些階段」可見。
    summary["figure_stages"] = dict(sorted(figure_stats.items()))
    # 快取採用各階段的件數，讓「哪個階段被選為最新版」可見（規格「可見性」節）。
    cache_stages: dict[str, int] = {}
    for _, kind, _ in latest_stage_map.values():
        cache_stages[kind] = cache_stages.get(kind, 0) + 1
    summary["figure_cache_stages"] = dict(sorted(cache_stages.items()))
    summary["figure_warnings"] = figure_warnings
    # 報表定案 #3 接線：已知 WIPS code 的新名稱變體即時補入唯一對照表（unknown/conflicting 進 manual）。
    summary["alias_variants"] = register_known_code_variants(
        variant_pairs, source_label=f"import:{path.name}",
        standardized_names=standardized_by_code)
    # 🔴 欄名不認得的警告（2026-08-03）：把「掃不到」與「沒有新的」分開。
    # ⚠ 沒有這個訊號時，兩者在 summary 上長得一模一樣（都是 alias_variants: 0），
    # 而前者是**整條名稱治理管線失效**、後者是正常狀態。
    if people_unmatched:
        summary["people_columns_unmatched"] = people_unmatched
        summary.setdefault("warnings", []).append(
            f"{people_unmatched} 列有申請人／專利權人資料，卻抽不出任何 (代碼, 名稱) 配對"
            "——欄名可能不在 PEOPLE_FIELD_COLUMNS 對照表內，"
            "公司名治理（自動建組、變體註冊、待中文化偵測）對這些列不會運作。"
        )
    return summary


def find_existing_raw_import(cur, file_hash: str) -> bool:
    """同 source_system＋source_file_hash 是否已匯入過。

    0019 後 source_files 表移除、來源 metadata（source_system/source_file_hash/imported_at）
    併入 raw_records；整檔冪等改查 raw_records。
    """
    cur.execute(
        "SELECT 1 FROM raw_records WHERE source_system = %s AND source_file_hash = %s LIMIT 1",
        (SOURCE_SYSTEM, file_hash),
    )
    return cur.fetchone() is not None


def insert_raw_record(cur, sheet_name: str, raw: dict[str, Any], file_hash: str) -> int:
    """寫 raw_records；來源 metadata 直接落本表（0019 後 schema）。"""
    # 圖片位元組不進 raw_data（JSONB 存不下也不該存）；追溯靠 patents."主附圖"。
    raw_json = {
        key: value_to_text(value)
        for key, value in raw.items()
        if key not in _INTERNAL_RECORD_KEYS
    }
    cur.execute(
        """
        INSERT INTO raw_records
            (sheet_name, row_number, raw_data, source_system, source_file_hash, imported_at)
        VALUES (%s, %s, %s::jsonb, %s, %s, now())
        RETURNING id
        """,
        (sheet_name, raw["_row_number"], json.dumps(raw_json, ensure_ascii=False),
         SOURCE_SYSTEM, file_hash),
    )
    return cur.fetchone()[0]


def upsert_patent(
    cur,
    patent: dict[str, Any],
    raw_record_id: int,
    stats: dict[str, int] | None = None,
    legal_status_conflicts: list[dict[str, Any]] | None = None,
) -> int:
    """依識別號 upsert 一筆 patent；命中既有走 update、否則 insert。

    傳入 stats 時累加：insert 路徑 stats["inserted"]；識別號命中既有一律 stats["matched_existing"]，
    其中真的有任一欄差異更新（update_patent_changed_fields 回 True）才另計 stats["updated"]。
    更新採 2026-07-22「差異即更新（新值非空）」政策，不改既有 mapping 與去重機制。
    """
    patent_params = build_patent_params(patent)
    patent_id = find_existing_patent_id(cur, patent)
    if patent_id:
        changed = update_patent_changed_fields(cur, patent_id, patent_params)
        if stats is not None:
            stats["matched_existing"] = stats.get("matched_existing", 0) + 1
            if changed:
                stats["updated"] = stats.get("updated", 0) + 1
        # 🔴 護欄二可見性（2026-08-07）：檔案帶了狀態值、但該案有人工登錄
        # 且值不同＝覆蓋被擋——記進衝突清單交使用者裁決，不得靜默吞掉。
        if legal_status_conflicts is not None:
            file_status = (patent_params.get("legal_status") or "").strip()
            if file_status:
                cur.execute(
                    "SELECT legal_status FROM patents "
                    f"WHERE id = %s AND {_MANUAL_STATUS_HISTORY_EXISTS} "
                    "AND legal_status IS DISTINCT FROM %s",
                    (patent_id, file_status),
                )
                row = cur.fetchone()
                if row is not None:
                    legal_status_conflicts.append({
                        "patent_id": int(patent_id),
                        "label": _patent_label(patent, patent_id),
                        "file_value": file_status,
                        "current_value": row[0],
                    })
        return patent_id

    # SQL 由 _UPDATE_COLUMN_PARAMS 生成（0046 起 insert／update 共用同一份欄位清單）
    cur.execute(
        _patent_insert_sql(),
        patent_params,
    )
    patent_id = cur.fetchone()[0]
    if stats is not None:
        stats["inserted"] = stats.get("inserted", 0) + 1
    return patent_id


def find_existing_patent_id(cur, patent: dict[str, Any]) -> int | None:
    for column_name in PATENT_IDENTIFIER_LOOKUP_ORDER:
        value = patent.get(column_name)
        if not value:
            continue
        cur.execute(
            f"""
            SELECT id
            FROM patents
            WHERE "{column_name}" = %s
            ORDER BY id
            LIMIT 1
            """,
            (value,),
        )
        existing = cur.fetchone()
        if existing:
            return existing[0]

    application_number = patent.get(APPLICATION_NUMBER_TRANSFORMED)
    if not application_number:
        return None
    cur.execute(
        """
        SELECT id
        FROM patents
        WHERE "申請號(轉換後)" = %s
          AND (%s::text IS NULL OR country_code IS NULL OR country_code = %s)
          AND (%s::text IS NULL OR database_name IS NULL OR database_name = %s)
        ORDER BY
            CASE WHEN country_code = %s THEN 0 ELSE 1 END,
            CASE WHEN database_name = %s THEN 0 ELSE 1 END,
            id
        LIMIT 1
        """,
        (
            application_number,
            patent.get("country_code"),
            patent.get("country_code"),
            patent.get("database_name"),
            patent.get("database_name"),
            patent.get("country_code"),
            patent.get("database_name"),
        ),
    )
    existing = cur.fetchone()
    return existing[0] if existing else None


# 差異比對用的參數名後綴：同一新值同時餵兩個 placeholder——寫回用原名（欄位型別語境，
# 由 COALESCE(欄) 錨定推型別）、文字比對用 <param>__cmp（一律 ::text 語境）。psycopg3
# server-side binding 對每個 placeholder 只推一種型別，拆成兩個名字才不會讓同一參數
# 同時被要求是 text 又是 date/int 而衝突（DatatypeMismatch / UndefinedFunction）。
_CHANGE_CMP_SUFFIX = "__cmp"

# 人工登錄判定（2026-08-07 護欄二）：人工線的歷程項目沒有 'source' 鍵、
# 匯入寫的一律帶 source='import'——任何一筆無 source 項目＝有人工判斷在。
_MANUAL_STATUS_HISTORY_EXISTS = (
    "EXISTS (SELECT 1 FROM jsonb_array_elements("
    "COALESCE(legal_status_history, '[]'::jsonb)) AS h(entry) "
    "WHERE NOT (h.entry ? 'source'))"
)


def _column_change_condition(column: str, param: str) -> str:
    """單欄「新值非空且與舊值不同」的判定 SQL（差異即更新政策的原子條件）。

    新值（<param>__cmp）與舊值一律 `::text` 後比較：①以 NULLIF(BTRIM(新值::text),'') 判空
    （NULL/空字串/全空白視為空），空則不觸發更新；②文字化兩側避免 date/int 欄與參數的
    型別不符。非空時以 IS DISTINCT FROM 比對（NULL 舊值也算差異）。此處只讀比對用參數，
    寫回值另由 SET 的原名參數處理，不受文字化影響。
    """
    cmp_param = f"{param}{_CHANGE_CMP_SUFFIX}"
    return (
        f"(NULLIF(BTRIM(%({cmp_param})s::text), '') IS NOT NULL "
        f"AND %({cmp_param})s::text IS DISTINCT FROM {column}::text)"
    )


def update_patent_changed_fields(cur, patent_id: int, patent_params: dict[str, Any]) -> bool:
    """逐欄比對：新值非空且與舊值不同就更新；新值空一律不覆蓋既有值。

    2026-07-22 政策（取代舊「只補 NULL」COALESCE 語意）：專利狀態會演進
    （legal_status、公告號、權利人、同族等），故命中既有時全欄一致採「差異即更新」。
    護欄：新值為空（NULL/空字串/全空白，NULLIF(BTRIM(...),'') 判定）不得覆蓋既有有值，
    避免某批來源缺欄把既有好資料清空。

    SET 每欄形態：欄 = CASE WHEN <新值非空且與舊值不同> THEN 新值 ELSE 欄 END；
    WHERE guard 為所有欄變更條件的 OR，讓完全無差異時 UPDATE 命中 0 列。
    回傳是否真的更新至少一欄（cur.rowcount>0），供匯入統計區分 matched_existing 與 updated
    （no-op 不算 updated）。SET/guard 皆由 _UPDATE_COLUMN_PARAMS 生成，不寫死欄名清單。
    """
    # THEN 用 COALESCE(%(param)s, 欄) 讓無型別參數由欄位錨定推得型別（沿用舊碼 COALESCE 錨定
    # 手法，date/int/text 皆適用）；THEN 只在新值非空時觸發，此時 COALESCE 必回新值本身。
    set_parts = []
    guard_parts = []
    for column, param in _UPDATE_COLUMN_PARAMS:
        cond = _column_change_condition(column, param)
        if column == "legal_status":
            # 🔴 護欄二（2026-08-07）：歷程含**人工登錄**（無 'source' 鍵的項目）
            # 的案，匯入不得靜默覆蓋——被擋下者由 upsert_patent 記進衝突清單。
            cond = f"({cond} AND NOT {_MANUAL_STATUS_HISTORY_EXISTS})"
            # 🔴 護欄三：匯入改狀態也 append 歷程（source: import），不斷鏈。
            set_parts.append(
                f"legal_status_history = CASE WHEN {cond} "
                "THEN COALESCE(legal_status_history, '[]'::jsonb) || "
                "jsonb_build_array(jsonb_build_object("
                "'from_status', legal_status, 'to_status', %(legal_status)s::text, "
                "'changed_at', to_jsonb(now()), 'source', 'import')) "
                "ELSE legal_status_history END"
            )
        set_parts.append(
            f"{column} = CASE WHEN {cond} "
            f"THEN COALESCE(%({param})s, {column}) ELSE {column} END"
        )
        guard_parts.append(cond)
    set_clause = ",\n            ".join(set_parts)
    guard = " OR ".join(guard_parts)
    # 每個新值同時綁定原名（寫回）與 <name>__cmp（文字比對），兩者取同一值；缺鍵者當 None。
    params: dict[str, Any] = {"patent_id": patent_id}
    for _, param in _UPDATE_COLUMN_PARAMS:
        value = patent_params.get(param)
        params[param] = value
        params[f"{param}{_CHANGE_CMP_SUFFIX}"] = value
    cur.execute(
        f"""
        UPDATE patents
        SET
            {set_clause}
        WHERE id = %(patent_id)s AND ({guard})
        """,
        params,
    )
    return cur.rowcount > 0


def _patent_label(patent: dict[str, Any], patent_id: int) -> str:
    """警告訊息用的專利識別標籤：優先取專利號，皆缺時退回內部 id。

    不寫死單一欄——依既有 PATENT_IDENTIFIER_LOOKUP_ORDER 與申請號依序取第一個有值者。
    """
    for column_name in (*PATENT_IDENTIFIER_LOOKUP_ORDER, "申請號"):
        value = patent.get(column_name)
        if value:
            return str(value)
    return f"id={patent_id}"


def update_patent_figures(cur, triplets: list[tuple[int, str, bytes]]) -> int:
    """批次寫入代表圖到 core_layer.patent_figures（一對多保存），回寫入列數。

    0031 起同一專利的各文獻階段（A 早期公開／B 審定公告／未知階段）各存一列，不再互相覆蓋。
    主鍵 (patent_id, document_kind) 天然去重：重匯同階段走 ON CONFLICT DO UPDATE 覆蓋內容。

    效率：一次 executemany（單次 round-trip 批送）寫完整批，不逐張發 INSERT
    （1900 筆時 N+1 會多出 1900 次往返）。空清單直接回 0、不發查詢。

    語意：triplets 由呼叫端只放「該列真的有圖」者（無圖列根本不入列），且同鍵重複已在
    呼叫端過濾並記入 figure_warnings，故此處不需再判空或去重。
    """
    if not triplets:
        return 0
    cur.executemany(
        """
        INSERT INTO patent_figures (patent_id, document_kind, content)
        VALUES (%s, %s, %s)
        ON CONFLICT (patent_id, document_kind) DO UPDATE
        SET content = EXCLUDED.content
        """,
        triplets,
    )
    return len(triplets)


def update_patent_figure_cache(cur, latest: dict[int, tuple[int, str, bytes]]) -> int:
    """批次回寫主表 patents."主附圖" 最新版快取，回更新列數。

    latest 為 patent_id → (rank, document_kind, 圖片 bytes)，由呼叫端以 figure_kind_rank
    取全批 rank 最大者（B > A > 未知階段）決定，故此處只負責批次寫入，不重做階段判定。
    前端與 API 讀主表即得最新版，無需 JOIN patent_figures（規格「主表欄位的定位」）。

    效率：單次 executemany 批送，不逐列 UPDATE；空 dict 直接回 0、不發查詢。
    """
    if not latest:
        return 0
    cur.executemany(
        f'UPDATE patents SET "{FIGURE_COLUMN}" = %s WHERE id = %s',
        [(blob, patent_id) for patent_id, (_, _, blob) in latest.items()],
    )
    return len(latest)


def insert_patent_source(cur, patent_id: int, raw_record_id: int) -> None:
    """寫入專利↔raw_record 來源對應。

    0021 patent_sources 僅 (patent_id, raw_record_id)（source_file_id 已隨 source_files 移除）；
    (patent_id, raw_record_id) 為主鍵，重覆對應以 ON CONFLICT DO NOTHING 略過。
    """
    cur.execute(
        """
        INSERT INTO patent_sources (patent_id, raw_record_id)
        VALUES (%s, %s)
        ON CONFLICT (patent_id, raw_record_id) DO NOTHING
        """,
        (patent_id, raw_record_id),
    )


def replace_people(cur, patent_id: int, raw_record_id: int, people: dict[str, Any]) -> None:
    """upsert 一列 patent_people；命中既有時採 2026-07-22「差異即更新（新值非空）」政策。

    盤點揪出的漏網：舊政策 COALESCE(EXCLUDED, 舊) 只補空，權利人/申請人演進（decision
    動因明列權利人）更新不到，且空白字串會覆蓋既有值。改為每欄
    CASE WHEN 新值非空（NULLIF(BTRIM,'') 判定）且與舊值不同 THEN 新值 ELSE 舊值 END，
    與 update_patent_changed_fields 同語意。people 欄皆 text，無需型別錨定。
    欄清單由 PEOPLE_FIELDS/PEOPLE_FIELD_COLUMNS 生成，不寫死欄名。
    """
    quoted_columns = ",\n            ".join(f'"{PEOPLE_FIELD_COLUMNS[field]}"' for field in PEOPLE_FIELDS)
    placeholders = ", ".join(["%s"] * (len(PEOPLE_FIELDS) + 1))
    update_assignments = ",\n            ".join(
        _people_column_update_clause(PEOPLE_FIELD_COLUMNS[field]) for field in PEOPLE_FIELDS
    )
    cur.execute(
        f"""
        INSERT INTO patent_people (
            patent_id,
            {quoted_columns}
        )
        VALUES ({placeholders})
        ON CONFLICT (patent_id) DO UPDATE
        SET
            {update_assignments}
        """,
        (patent_id, *(people.get(field) for field in PEOPLE_FIELDS)),
    )


def _people_column_update_clause(column: str) -> str:
    """單一 people 欄的 ON CONFLICT DO UPDATE 子句（差異即更新、新值空不覆蓋）。

    EXCLUDED.<欄> 為本次匯入新值、patent_people.<欄> 為既有值。新值非空
    （NULLIF(BTRIM(...),'') 判定）且與舊值不同才寫入新值，否則保留舊值。people 欄皆 text，
    直接比較不需型別錨定。
    """
    excluded = f'EXCLUDED."{column}"'
    existing = f'patent_people."{column}"'
    return (
        f'"{column}" = CASE '
        f"WHEN NULLIF(BTRIM({excluded}), '') IS NOT NULL "
        f"AND {excluded} IS DISTINCT FROM {existing} "
        f"THEN {excluded} ELSE {existing} END"
    )


def replace_attributes(cur, patent_id: int, raw_record_id: int, attributes: dict[str, Any]) -> None:
    cur.execute("DELETE FROM patent_attributes WHERE patent_id = %s AND raw_record_id = %s", (patent_id, raw_record_id))
    quoted_columns = ",\n            ".join(f'"{column}"' for column in ATTRIBUTE_FIELD_COLUMNS.values())
    placeholders = ", ".join(["%s"] * (len(ATTRIBUTE_FIELDS) + 2))
    cur.execute(
        f"""
        INSERT INTO patent_attributes (
            patent_id, raw_record_id,
            {quoted_columns}
        )
        VALUES ({placeholders})
        """,
        (
            patent_id,
            raw_record_id,
            *(attributes.get(field) for field in ATTRIBUTE_FIELDS),
        ),
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Import WIPS XLSX into patent_ppt database.")
    parser.add_argument("path", type=Path, help="Path to WIPS XLSX file.")
    parser.add_argument("--dry-run", action="store_true", help="Read and normalize without writing to database.")
    args = parser.parse_args()
    summary = import_wips_file(args.path, dry_run=args.dry_run)
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
